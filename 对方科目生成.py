# -*- coding: utf-8 -*-
# 对方科目生成工具 v1.8.0
# 将序时账数据自动生成对方科目，支持多核并行计算和异常分录检测
# v1.8.0: 核心算法重构为折半枚举（MITM），解决长尾凑单超时问题
# v1.7.0: 新增多公司账套支持，按账套+月+凭证号分组

import sys
import time
import datetime
import functools
import bisect
import os
import math
import traceback
import multiprocessing
import threading
import queue
from concurrent.futures import ProcessPoolExecutor, as_completed
from typing import List, Dict, Any, Optional, Tuple, Set, Union
from decimal import Decimal, getcontext, ROUND_HALF_EVEN, InvalidOperation
import pandas as pd

try:
    from openpyxl.chart import LineChart, Reference
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    LineChart = Reference = get_column_letter = None

import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk

try:
    import customtkinter as ctk
    USE_CUSTOMTKINTER = True  # 使用CustomTkinter增强版界面
except ImportError:
    USE_CUSTOMTKINTER = False
    ctk = None

# --- 全局配置与优化 ---
# 提高递归限制以支持复杂凭证的深度优先搜索（DFS）算法
sys.setrecursionlimit(20000)

# 常量定义
MAX_SUBSET_ITERATIONS = 1000000  # 子集和算法最大迭代次数（旧DFS算法保留）
PROGRESS_UPDATE_FREQ = 0.5       # 进度条刷新频率(秒)
TOLERANCE = 0.005                # 金额匹配容差
FLOAT_EPSILON = 1e-9             # 浮点数相等判断阈值
AMOUNT_VALID_THRESHOLD = 1e-6    # 发生额有效值判断阈值
AMOUNT_CHECK_THRESHOLD = 0.01    # 金额校验阈值
REQUIRED_OUTPUT_COLUMNS = ['借方发生额', '贷方发生额', '对方科目']
DEFAULT_CPU_CORES = 4            # CPU核心数默认值

# 折半枚举算法（MITM）安全常量 - 取值依据说明：
# MAX_CANDIDATE_POOL_SIZE = 36: 单边18个元素，最大组合数2^18≈26万，计算可控
MAX_CANDIDATE_POOL_SIZE = 36
# MITM_ITERATION_LIMIT = 300000: 略大于2^18，作为兜底保护
MITM_ITERATION_LIMIT = 300000

# 账套列识别关键词（用于多公司账套检测）
LEDGER_ACCOUNT_KEYWORDS = ['账套', '核算账套', '公司账套', '账套名', '公司名称', '核算主体', '主体名称']

# ==========================================
# 金融级精度引擎
# ==========================================
getcontext().prec = 28

class PrecisionEngine:
    """
    金融级精度处理引擎。
    
    核心原理：
    1. 使用Decimal替代float进行精确计算
    2. 转换为整数"厘"进行比较，避免浮点误差
    3. 采用银行家舍入（四舍六入五成双）
    4. 最终输出保留2位小数
    """
    SCALE = 10000  # 放大倍数：元 -> 厘（0.0001元）
    DECIMAL_QUANTIZER = Decimal("0.01")  # 输出量化器：2位小数
    LI_QUANTIZER = Decimal("0.0001")     # 内部计算量化器：4位小数（厘精度）
    LI_TOLERANCE = 100  # 整数厘容差（0.01元 = 1分）
    
    @staticmethod
    def to_decimal(value: Any) -> Decimal:
        """
        将任意数值转换为精确的Decimal（2位小数，用于输出）。
        
        :param value: 输入值（float/int/str/Decimal）
        :return: 量化后的Decimal（2位小数，银行家舍入）
        """
        if value is None:
            return Decimal("0")
        try:
            if isinstance(value, Decimal):
                d = value
            elif isinstance(value, float):
                d = Decimal(str(value))
            else:
                d = Decimal(str(value))
            return d.quantize(PrecisionEngine.DECIMAL_QUANTIZER, rounding=ROUND_HALF_EVEN)
        except (InvalidOperation, ValueError, TypeError):
            return Decimal("0")
    
    @staticmethod
    def to_integer_li(value: Any) -> int:
        """
        将金额转换为整数厘（用于精确计算和比较）。
        
        :param value: 金额值（元）
        :return: 整数厘（元 × 10000）
        """
        if value is None:
            return 0
        try:
            if isinstance(value, Decimal):
                d = value
            elif isinstance(value, float):
                d = Decimal(str(value))
            else:
                d = Decimal(str(value))
            # 先量化到厘精度（4位小数），再转为整数
            d_li = d.quantize(PrecisionEngine.LI_QUANTIZER, rounding=ROUND_HALF_EVEN)
            return int(d_li * PrecisionEngine.SCALE)
        except (InvalidOperation, ValueError, TypeError):
            return 0
    
    @staticmethod
    def from_integer_li(li: int) -> Decimal:
        """
        从整数厘转换为Decimal元（2位小数）。
        
        :param li: 整数厘
        :return: Decimal元（2位小数）
        """
        result = Decimal(li) / PrecisionEngine.SCALE
        return result.quantize(PrecisionEngine.DECIMAL_QUANTIZER, rounding=ROUND_HALF_EVEN)
    
    @staticmethod
    def amounts_match(amount1: Any, amount2: Any, tolerance_li: int = None) -> bool:
        """
        判断两个金额是否在容差范围内匹配。
        
        :param amount1: 金额1
        :param amount2: 金额2
        :param tolerance_li: 容差（整数厘），默认100厘=0.01元=1分
        :return: 是否匹配
        """
        if tolerance_li is None:
            tolerance_li = PrecisionEngine.LI_TOLERANCE
        li1 = PrecisionEngine.to_integer_li(amount1)
        li2 = PrecisionEngine.to_integer_li(amount2)
        return abs(li1 - li2) <= tolerance_li
    
    @staticmethod
    def sum_amounts(amounts: List[Any]) -> Decimal:
        """
        精确求和。
        
        :param amounts: 金额列表
        :return: 精确总和（Decimal，2位小数）
        """
        total = Decimal("0")
        for amt in amounts:
            total += PrecisionEngine.to_decimal(amt)
        return total
    
    @staticmethod
    def compare_amounts(amount1: Any, amount2: Any) -> int:
        """
        比较两个金额大小。
        
        :param amount1: 金额1
        :param amount2: 金额2
        :return: -1/0/1（小于/等于/大于）
        """
        li1 = PrecisionEngine.to_integer_li(amount1)
        li2 = PrecisionEngine.to_integer_li(amount2)
        if li1 < li2:
            return -1
        elif li1 > li2:
            return 1
        return 0


# 兼容旧接口的别名
to_integer_cents = PrecisionEngine.to_integer_li
from_integer_cents = PrecisionEngine.from_integer_li


def amounts_match_precision(amount1: Any, amount2: Any, tolerance_li: int = 200) -> bool:
    """
    精确金额匹配函数（兼容旧接口）。
    
    :param amount1: 金额1
    :param amount2: 金额2
    :param tolerance_li: 容差（厘），默认200厘=0.02元=2分
    :return: 是否匹配
    """
    return PrecisionEngine.amounts_match(amount1, amount2, tolerance_li)


def sum_amounts_precision(amounts: List[Any]) -> Decimal:
    """
    精确求和函数（兼容旧接口）。
    
    :param amounts: 金额列表
    :return: 精确总和
    """
    return PrecisionEngine.sum_amounts(amounts)

# ==========================================
# 进度条与GUI组件
# ==========================================

class TerminalProgressBar:
    """终端进度条组件 (简化同步版)。"""
    def __init__(self, update_freq: float = 0.5):
        self.update_freq = update_freq
        self.last_update_time = 0.0
        
    def update(self, percent: float, message: str, phase: Optional[str] = None):
        """更新进度状态。"""
        now = time.time()
        if now - self.last_update_time < self.update_freq and percent < 100:
            return

        self.last_update_time = now
        self._print_line(percent, message, phase)
        
        if percent >= 100:
            sys.stdout.write("\n")
            sys.stdout.flush()

    def stop(self):
        """停止/完成。"""
        sys.stdout.write("\n")

    def _print_line(self, percent: float, message: str, phase: Optional[str]):
        """打印单行进度条。"""
        bar_len = 30
        filled = int(bar_len * percent / 100)
        bar = '█' * filled + '░' * (bar_len - filled)
        phase_str = f"[{phase}]" if phase else ""
        
        status_line = (
            f"\r{phase_str:<15} "
            f"{bar} "
            f"{percent:>5.1f}% | "
            f"{message:<20}"
        )
        try:
            sys.stdout.write(status_line)
            sys.stdout.flush()
        except Exception:
            pass

TERM_BAR = TerminalProgressBar(update_freq=PROGRESS_UPDATE_FREQ)

class CTkProgressBar:
    """CustomTkinter 进度条组件。"""
    def __init__(self, parent=None):
        self.parent = parent
        self.window = None
        self.progress = None
        self.label = None
        self.phase_label = None

    def show(self, title: str = "处理中"):
        """显示进度条窗口。"""
        if not USE_CUSTOMTKINTER:
            return

        self.window = ctk.CTkToplevel() if self.parent else ctk.CTk()
        self.window.title(title)
        self.window.geometry("500x150")
        if self.parent:
            self.window.transient(self.parent)
        self.window.grab_set()

        self.phase_label = ctk.CTkLabel(self.window, text="准备中...", font=("微软雅黑", 12))
        self.phase_label.pack(pady=(20, 5))

        self.progress = ctk.CTkProgressBar(self.window, width=400)
        self.progress.pack(pady=10)
        self.progress.set(0)

        self.label = ctk.CTkLabel(self.window, text="正在初始化...", font=("微软雅黑", 10))
        self.label.pack(pady=5)

        self.window.update()

    def update(self, percent: float, message: str, phase: Optional[str] = None):
        """更新进度。"""
        if self.window and USE_CUSTOMTKINTER:
            if phase:
                self.phase_label.configure(text=phase)
            self.label.configure(text=message)
            self.progress.set(percent / 100.0)
            self.window.update()

    def close(self):
        """关闭进度条。"""
        if self.window:
            self.window.destroy()
            self.window = None

class CustomInputDialog:
    """自定义输入对话框，支持数字输入。"""
    def __init__(self, parent, title: str, prompt: str, initial_value: float = 0.0, min_value: float = None, max_value: float = None):
        self.result = None
        self.window = None
        
        if USE_CUSTOMTKINTER:
            self._create_ctk_dialog(parent, title, prompt, initial_value, min_value, max_value)
        else:
            self._create_tk_dialog(parent, title, prompt, initial_value, min_value, max_value)

    def _create_ctk_dialog(self, parent, title, prompt, initial_value, min_value, max_value):
        self.window = ctk.CTkToplevel() if parent else ctk.CTk()
        self.window.title(title)
        self.window.geometry("400x200")
        if parent:
            self.window.transient(parent)
        self.window.grab_set()

        ctk.CTkLabel(self.window, text=prompt, font=("微软雅黑", 12)).pack(pady=20)
        self.entry = ctk.CTkEntry(self.window, width=250, font=("微软雅黑", 12))
        self.entry.insert(0, str(initial_value))
        self.entry.pack(pady=10)
        self._add_hint(self.window, min_value, max_value, is_ctk=True)
        self._add_buttons(self.window, is_ctk=True)
        
        self.entry.focus()
        self.entry.bind("<Return>", lambda e: self._on_ok())
        self.window.wait_window()

    def _create_tk_dialog(self, parent, title, prompt, initial_value, min_value, max_value):
        self.window = tk.Toplevel(parent) if parent else tk.Tk()
        self.window.title(title)
        self.window.geometry("400x200")
        if parent:
            self.window.transient(parent)
        self.window.grab_set()

        tk.Label(self.window, text=prompt, font=("微软雅黑", 12)).pack(pady=20)
        self.entry = tk.Entry(self.window, width=30, font=("微软雅黑", 12))
        self.entry.insert(0, str(initial_value))
        self.entry.pack(pady=10)
        self._add_hint(self.window, min_value, max_value, is_ctk=False)
        self._add_buttons(self.window, is_ctk=False)
        
        self.entry.focus()
        self.entry.bind("<Return>", lambda e: self._on_ok())
        self.window.wait_window()

    def _add_hint(self, window, min_value, max_value, is_ctk):
        if min_value is not None or max_value is not None:
            hint_text = ""
            if min_value is not None: hint_text += f"最小值: {min_value}"
            if max_value is not None: hint_text += f", 最大值: {max_value}"
            
            if is_ctk:
                ctk.CTkLabel(window, text=hint_text, font=("微软雅黑", 9), text_color="gray").pack(pady=5)
            else:
                tk.Label(window, text=hint_text, font=("微软雅黑", 9), fg="gray").pack(pady=5)

    def _add_buttons(self, window, is_ctk):
        if is_ctk:
            btn_frame = ctk.CTkFrame(window)
            btn_frame.pack(pady=20)
            ctk.CTkButton(btn_frame, text="确定", width=80, command=self._on_ok).pack(side="left", padx=10)
            ctk.CTkButton(btn_frame, text="取消", width=80, command=self._on_cancel).pack(side="left", padx=10)
        else:
            btn_frame = tk.Frame(window)
            btn_frame.pack(pady=20)
            tk.Button(btn_frame, text="确定", width=10, command=self._on_ok).pack(side="left", padx=10)
            tk.Button(btn_frame, text="取消", width=10, command=self._on_cancel).pack(side="left", padx=10)

    def _on_ok(self):
        try:
            value = float(self.entry.get())
            self.result = value
            self.window.destroy()
        except ValueError:
            messagebox.showwarning("输入错误", "请输入有效的数字！")

    def _on_cancel(self):
        self.result = None
        self.window.destroy()

    def get_result(self):
        return self.result

class CustomMessageBox:
    """自定义消息框。"""
    @staticmethod
    def _show_ctk_dialog(parent, title, message, kind):
        dialog = ctk.CTkToplevel(parent) if parent else ctk.CTk()
        dialog.title(title)
        dialog.geometry("420x180")
        if parent:
            dialog.transient(parent)
        dialog.grab_set()

        color = {"error": "#FF6B6B", "warning": "#FFB020"}.get(kind, "#3B8ED0")
        
        ctk.CTkLabel(dialog, text=title, font=("微软雅黑", 14, "bold"), text_color=color).pack(pady=(18, 8), padx=16)
        ctk.CTkLabel(dialog, text=message, font=("微软雅黑", 11), wraplength=380, justify="left").pack(pady=(0, 12), padx=16)
        
        btn = ctk.CTkButton(dialog, text="确定", width=90, command=dialog.destroy)
        btn.pack(pady=(0, 16))
        btn.focus_set()
        
        dialog.bind("<Return>", lambda _e: dialog.destroy())
        dialog.bind("<Escape>", lambda _e: dialog.destroy())
        dialog.wait_window()

    @staticmethod
    def showinfo(title, message, parent=None):
        if USE_CUSTOMTKINTER and parent is not None:
            CustomMessageBox._show_ctk_dialog(parent, title, message, kind="info")
        else:
            messagebox.showinfo(title, message)

    @staticmethod
    def showwarning(title, message, parent=None):
        if USE_CUSTOMTKINTER and parent is not None:
            CustomMessageBox._show_ctk_dialog(parent, title, message, kind="warning")
        else:
            messagebox.showwarning(title, message)

    @staticmethod
    def showerror(title, message, parent=None):
        if USE_CUSTOMTKINTER and parent is not None:
            CustomMessageBox._show_ctk_dialog(parent, title, message, kind="error")
        else:
            messagebox.showerror(title, message)

class GuiProgressManager:
    """GUI 进度条管理器 (线程安全)。"""
    def __init__(self):
        self.progress_bar = None
        self.use_gui = USE_CUSTOMTKINTER or tk
        self.msg_queue = queue.Queue()
        self._gui_callback = None

    def set_gui_callback(self, callback):
        """设置GUI回调函数，用于主线程更新。"""
        self._gui_callback = callback

    def create_progress_bar(self, parent=None):
        # 在主线程中创建，或者由外部管理
        pass

    def update(self, percent: float, message: str, phase: Optional[str] = None):
        """更新进度 (可在任意线程调用)。"""
        if self._gui_callback:
            self.msg_queue.put(("update", percent, message, phase))
            # 通知主线程有更新 (如果是通过轮询则不需要此步，这里假设主线程会轮询)
        else:
            TERM_BAR.update(percent, message, phase)

    def stop(self):
        """停止/完成。"""
        if self._gui_callback:
            self.msg_queue.put(("stop", None, None, None))
        else:
            TERM_BAR.stop()

    def close(self):
        self.stop()

GUI_PROGRESS = GuiProgressManager()

# ==========================================
# 常用会计分录规则库
# ==========================================

STANDARD_RULES = [
    {"credit": ["发出商品"], "debit": ["主营业务成本"]},
    {"credit": ["其他收益"], "debit": ["银行存款"]},
    {"credit": ["合同负债"], "debit": ["银行存款"]},
    {"credit": ["银行存款"], "debit": ["管理费用"]},
    {"credit": ["工程物资"], "debit": ["在建工程"]},
    {"credit": ["银行存款"], "debit": ["营业外收入"]},
    {"credit": ["应收账款"], "debit": ["财务费用"]},
    {"credit": ["实收资本"], "debit": ["无形资产"]},
    {"credit": ["应付职工薪酬"], "debit": ["研发支出"]},
    {"credit": ["累计摊销"], "debit": ["管理费用"]},
    {"credit": ["银行存款"], "debit": ["短期借款"]},
    {"credit": ["投资性房地产"], "debit": ["公允价值变动损益"]},
    {"credit": ["应付职工薪酬"], "debit": ["生产成本"]},
    {"credit": ["油气资产减值准备"], "debit": ["资产减值损失"]},
    {"credit": ["贷款损失准备"], "debit": ["信用减值损失"]},
    {"credit": ["固定资产清理"], "debit": ["银行存款"]},
    {"credit": ["应收账款"], "debit": ["应交税费"]},
    {"credit": ["银行存款"], "debit": ["在建工程"]},
    {"credit": ["累计摊销"], "debit": ["销售费用"]},
    {"credit": ["投资收益"], "debit": ["长期股权投资"]},
    {"credit": ["待处理财产损溢"], "debit": ["营业外支出"]},
    {"credit": ["应付账款"], "debit": ["管理费用"]},
    {"credit": ["应付账款"], "debit": ["营业外收入"]},
    {"credit": ["实收资本"], "debit": ["长期股权投资"]},
    {"credit": ["主营业务收入"], "debit": ["合同负债"]},
    {"credit": ["其他应付款"], "debit": ["应付职工薪酬"]},
    {"credit": ["其他权益工具投资"], "debit": ["投资收益"]},
    {"credit": ["银行存款"], "debit": ["营业外支出"]},
    {"credit": ["应收股利"], "debit": ["银行存款"]},
    {"credit": ["主营业务收入"], "debit": ["应收账款"]},
    {"credit": ["固定资产减值准备"], "debit": ["资产减值损失"]},
    {"credit": ["投资性房地产"], "debit": ["无形资产"]},
    {"credit": ["在建工程减值准备"], "debit": ["资产减值损失"]},
    {"credit": ["库存股"], "debit": ["盈余公积"]},
    {"credit": ["待处理财产损溢"], "debit": ["固定资产"]},
    {"credit": ["周转材料"], "debit": ["其他业务成本"]},
    {"credit": ["固定资产清理"], "debit": ["资产处置损益"]},
    {"credit": ["坏账准备"], "debit": ["资产减值损失"]},
    {"credit": ["银行存款"], "debit": ["固定资产"]},
    {"credit": ["长期股权投资"], "debit": ["应收股利"]},
    {"credit": ["应交税费"], "debit": ["应付职工薪酬"]},
    {"credit": ["固定资产"], "debit": ["投资性房地产"]},
    {"credit": ["持有待售资产减值准备"], "debit": ["资产减值损失"]},
    {"credit": ["银行存款"], "debit": ["应付职工薪酬"]},
    {"credit": ["投资收益"], "debit": ["资本公积"]},
    {"credit": ["应交税费"], "debit": ["税金及附加"]},
    {"credit": ["应付账款"], "debit": ["工程物资"]},
    {"credit": ["商誉减值准备"], "debit": ["资产减值损失"]},
    {"credit": ["其他权益工具投资"], "debit": ["银行存款"]},
    {"credit": ["无形资产"], "debit": ["累计摊销"]},
    {"credit": ["应收账款"], "debit": ["应收票据"]},
    {"credit": ["合同资产减值准备"], "debit": ["资产减值损失"]},
    {"credit": ["银行存款"], "debit": ["长期借款"]},
    {"credit": ["股本"], "debit": ["盈余公积"]},
    {"credit": ["应付账款"], "debit": ["固定资产"]},
    {"credit": ["应付利息"], "debit": ["财务费用"]},
    {"credit": ["其他应收款"], "debit": ["管理费用"]},
    {"credit": ["固定资产"], "debit": ["固定资产减值准备"]},
    {"credit": ["无形资产"], "debit": ["投资性房地产"]},
    {"credit": ["银行存款"], "debit": ["应付债券"]},
    {"credit": ["实收资本"], "debit": ["库存商品"]},
    {"credit": ["其他应收款"], "debit": ["销售费用"]},
    {"credit": ["投资收益"], "debit": ["应收股利"]},
    {"credit": ["银行存款"], "debit": ["交易性金融资产"]},
    {"credit": ["固定资产清理"], "debit": ["营业外收入"]},
    {"credit": ["应收账款"], "debit": ["主营业务收入"]},
    {"credit": ["原材料"], "debit": ["待处理财产损溢"]},
    {"credit": ["资本公积"], "debit": ["无形资产"]},
    {"credit": ["交易性金融资产"], "debit": ["公允价值变动损益"]},
    {"credit": ["累计摊销"], "debit": ["制造费用"]},
    {"credit": ["坏账准备"], "debit": ["信用减值损失"]},
    {"credit": ["投资收益"], "debit": ["应收利息"]},
    {"credit": ["投资性房地产减值准备"], "debit": ["资产减值损失"]},
    {"credit": ["交易性金融资产"], "debit": ["投资收益"]},
    {"credit": ["预付账款"], "debit": ["在途物资"]},
    {"credit": ["银行存款"], "debit": ["销售费用"]},
    {"credit": ["其他综合收益"], "debit": ["其他权益工具投资"]},
    {"credit": ["库存股"], "debit": ["应付职工薪酬"]},
    {"credit": ["银行存款"], "debit": ["其他债权投资"]},
    {"credit": ["财务费用"], "debit": ["银行存款"]},
    {"credit": ["长期待摊费用"], "debit": ["管理费用"]},
    {"credit": ["资本公积"], "debit": ["长期股权投资"]},
    {"credit": ["固定资产清理"], "debit": ["营业外支出"]},
    {"credit": ["投资收益"], "debit": ["银行存款"]},
    {"credit": ["主营业务收入"], "debit": ["银行存款"]},
    {"credit": ["长期股权投资减值准备"], "debit": ["资产减值损失"]},
    {"credit": ["递延所得税负债"], "debit": ["所得税费用"]},
    {"credit": ["实收资本"], "debit": ["银行存款"]},
    {"credit": ["长期股权投资"], "debit": ["持有待售资产"]},
    {"credit": ["其他债权投资"], "debit": ["投资收益"]},
    {"credit": ["应付账款"], "debit": ["销售费用"]},
    {"credit": ["银行存款"], "debit": ["研发支出"]},
    {"credit": ["原材料"], "debit": ["其他业务成本"]},
    {"credit": ["主营业务成本"], "debit": ["库存商品"]},
    {"credit": ["递延所得税资产"], "debit": ["所得税费用"]},
    {"credit": ["投资性房地产"], "debit": ["其他业务成本"]},
    {"credit": ["银行存款"], "debit": ["工程物资"]},
    {"credit": ["银行存款"], "debit": ["应付票据"]},
    {"credit": ["银行存款"], "debit": ["其他货币资金"]},
    {"credit": ["银行存款"], "debit": ["合同履约成本"]},
    {"credit": ["原材料"], "debit": ["在建工程"]},
    {"credit": ["实收资本"], "debit": ["原材料"]},
    {"credit": ["交易性金融资产"], "debit": ["银行存款"]},
    {"credit": ["累计折旧"], "debit": ["其他业务成本"]},
    {"credit": ["其他应收款"], "debit": ["库存现金"]},
    {"credit": ["债权投资"], "debit": ["银行存款"]},
    {"credit": ["预付账款"], "debit": ["原材料"]},
    {"credit": ["短期借款"], "debit": ["银行存款"]},
    {"credit": ["银行存款"], "debit": ["材料采购"]},
    {"credit": ["固定资产"], "debit": ["固定资产清理"]},
    {"credit": ["交易性金融资产减值准备"], "debit": ["资产减值损失"]},
    {"credit": ["主营业务收入"], "debit": ["合同资产"]},
    {"credit": ["其他应付款"], "debit": ["银行存款"]},
    {"credit": ["投资性房地产"], "debit": ["银行存款"]},
    {"credit": ["其他业务收入"], "debit": ["预收账款"]},
    {"credit": ["应收利息"], "debit": ["银行存款"]},
    {"credit": ["银行存款"], "debit": ["应付股利"]},
    {"credit": ["应付职工薪酬"], "debit": ["在建工程"]},
    {"credit": ["其他应收款"], "debit": ["制造费用"]},
    {"credit": ["其他债权投资"], "debit": ["银行存款"]},
    {"credit": ["应交税费"], "debit": ["预收账款"]},
    {"credit": ["库存商品"], "debit": ["待处理财产损溢"]},
    {"credit": ["银行存款"], "debit": ["库存现金"]},
    {"credit": ["应收票据"], "debit": ["银行存款"]},
    {"credit": ["主营业务收入"], "debit": ["预计负债"]},
    {"credit": ["应付账款"], "debit": ["材料采购"]},
    {"credit": ["银行存款"], "debit": ["投资性房地产"]},
    {"credit": ["银行存款"], "debit": ["其他收益"]},
    {"credit": ["长期股权投资"], "debit": ["其他综合收益"]},
    {"credit": ["研发支出"], "debit": ["无形资产"]},
    {"credit": ["待处理财产损溢"], "debit": ["以前年度损益调整"]},
    {"credit": ["应交税费"], "debit": ["以前年度损益调整"]},
    {"credit": ["固定资产"], "debit": ["累计折旧"]},
    {"credit": ["主营业务收入"], "debit": ["合同结算"]},
    {"credit": ["银行存款"], "debit": ["无形资产"]},
    {"credit": ["银行存款"], "debit": ["以前年度损益调整"]},
    {"credit": ["银行存款"], "debit": ["制造费用"]},
    {"credit": ["投资性房地产"], "debit": ["资产处置损益"]},
    {"credit": ["库存股"], "debit": ["资本公积"]},
    {"credit": ["应付账款"], "debit": ["投资性房地产"]},
    {"credit": ["无形资产"], "debit": ["银行存款"]},
    {"credit": ["应付账款"], "debit": ["在途物资"]},
    {"credit": ["银行存款"], "debit": ["长期股权投资"]},
    {"credit": ["银行存款"], "debit": ["应付利息"]},
    {"credit": ["其他业务收入"], "debit": ["应收账款"]},
    {"credit": ["银行存款"], "debit": ["合同负债"]},
    {"credit": ["应付账款"], "debit": ["无形资产"]},
    {"credit": ["待处理财产损溢"], "debit": ["管理费用"]},
    {"credit": ["应交税费"], "debit": ["递延所得税负债"]},
    {"credit": ["资本公积"], "debit": ["银行存款"]},
    {"credit": ["应付账款"], "debit": ["制造费用"]},
    {"credit": ["材料采购"], "debit": ["原材料"]},
    {"credit": ["银行存款"], "debit": ["其他业务收入"]},
    {"credit": ["应交税费"], "debit": ["应收账款"]},
    {"credit": ["营业外收入"], "debit": ["银行存款"]},
    {"credit": ["银行存款"], "debit": ["其他应收款"]},
    {"credit": ["固定资产"], "debit": ["待处理财产损溢"]},
    {"credit": ["在建工程"], "debit": ["固定资产"]},
    {"credit": ["银行存款"], "debit": ["其他权益工具投资"]},
    {"credit": ["生产性生物资产减值准备"], "debit": ["资产减值损失"]},
    {"credit": ["无形资产"], "debit": ["无形资产减值准备"]},
    {"credit": ["累计折旧"], "debit": ["管理费用"]},
    {"credit": ["应收票据"], "debit": ["财务费用"]},
    {"credit": ["应收账款"], "debit": ["其他业务收入"]},
    {"credit": ["主营业务成本"], "debit": ["预计负债"]},
    {"credit": ["无形资产"], "debit": ["资产处置损益"]},
    {"credit": ["主营业务收入"], "debit": ["应收票据"]},
    {"credit": ["累计折旧"], "debit": ["销售费用"]},
    {"credit": ["应付股利"], "debit": ["股本"]},
    {"credit": ["实收资本"], "debit": ["固定资产"]},
    {"credit": ["银行存款"], "debit": ["库存股"]},
    {"credit": ["应付职工薪酬"], "debit": ["管理费用"]},
    {"credit": ["营业外收入"], "debit": ["待处理财产损溢"]},
    {"credit": ["银行存款"], "debit": ["商誉"]},
    {"credit": ["库存商品"], "debit": ["发出商品"]},
    {"credit": ["银行存款"], "debit": ["投资收益"]},
    {"credit": ["其他业务收入"], "debit": ["应交税费"]},
    {"credit": ["应付票据"], "debit": ["应付账款"]},
    {"credit": ["持有待售资产"], "debit": ["资产处置损益"]},
    {"credit": ["待处理财产损溢"], "debit": ["库存商品"]},
    {"credit": ["所得税费用"], "debit": ["递延所得税负债"]},
    {"credit": ["制造费用"], "debit": ["生产成本"]},
    {"credit": ["其他综合收益"], "debit": ["其他债权投资"]},
    {"credit": ["应交税费"], "debit": ["固定资产清理"]},
    {"credit": ["银行存款"], "debit": ["应交税费"]},
    {"credit": ["材料采购"], "debit": ["材料成本差异"]},
    {"credit": ["原材料"], "debit": ["合同履约成本"]},
    {"credit": ["银行存款"], "debit": ["固定资产清理"]},
    {"credit": ["股本"], "debit": ["长期股权投资"]},
    {"credit": ["无形资产"], "debit": ["营业外收入"]},
    {"credit": ["其他应收款"], "debit": ["银行存款"]},
    {"credit": ["应交税费"], "debit": ["递延所得税资产"]},
    {"credit": ["银行存款"], "debit": ["在途物资"]},
    {"credit": ["投资收益"], "debit": ["交易性金融资产"]},
    {"credit": ["固定资产"], "debit": ["营业外支出"]},
    {"credit": ["工程物资减值准备"], "debit": ["资产减值损失"]},
    {"credit": ["其他业务收入"], "debit": ["银行存款"]},
    {"credit": ["投资性房地产"], "debit": ["固定资产"]},
    {"credit": ["应付账款"], "debit": ["应交税费"]},
    {"credit": ["无形资产"], "debit": ["在建工程"]},
    {"credit": ["累计摊销"], "debit": ["其他业务成本"]},
    {"credit": ["应付职工薪酬"], "debit": ["合同履约成本"]},
    {"credit": ["应交税费"], "debit": ["银行存款"]},
    {"credit": ["合同履约成本"], "debit": ["主营业务成本"]},
    {"credit": ["债权投资减值准备"], "debit": ["信用减值损失"]},
    {"credit": ["应交税费"], "debit": ["所得税费用"]},
    {"credit": ["无形资产"], "debit": ["营业外支出"]},
    {"credit": ["其他债权投资减值准备"], "debit": ["信用减值损失"]},
    {"credit": ["应交税费"], "debit": ["其他应付款"]},
    {"credit": ["银行存款"], "debit": ["预付账款"]},
    {"credit": ["银行存款"], "debit": ["其他应付款"]},
    {"credit": ["库存现金"], "debit": ["银行存款"]},
    {"credit": ["投资收益"], "debit": ["其他综合收益"]},
    {"credit": ["库存股"], "debit": ["股本"]},
    {"credit": ["存货跌价准备"], "debit": ["资产减值损失"]},
    {"credit": ["其他应收款"], "debit": ["坏账准备"]},
    {"credit": ["长期股权投资"], "debit": ["投资收益"]},
    {"credit": ["应交税费"], "debit": ["应收票据"]},
    {"credit": ["银行存款"], "debit": ["长期待摊费用"]},
    {"credit": ["应收账款"], "debit": ["银行存款"]},
    {"credit": ["原材料"], "debit": ["制造费用"]},
    {"credit": ["待处理财产损溢"], "debit": ["原材料"]},
    {"credit": ["债权投资"], "debit": ["投资收益"]},
    {"credit": ["股本"], "debit": ["资本公积"]},
    {"credit": ["库存商品"], "debit": ["主营业务成本"]},
    {"credit": ["累计折旧"], "debit": ["制造费用"]},
    {"credit": ["固定资产"], "debit": ["持有待售资产"]},
    {"credit": ["所得税费用"], "debit": ["递延所得税资产"]},
    {"credit": ["银行存款"], "debit": ["原材料"]},
    {"credit": ["生产成本"], "debit": ["库存商品"]},
    {"credit": ["以前年度损益调整"], "debit": ["银行存款"]},
    {"credit": ["在途物资"], "debit": ["原材料"]},
    {"credit": ["资产减值损失"], "debit": ["存货跌价准备"]},
    {"credit": ["递延收益"], "debit": ["其他收益"]},
    {"credit": ["资本公积"], "debit": ["固定资产"]},
    {"credit": ["发出商品"], "debit": ["应收账款"]},
    {"credit": ["应付职工薪酬"], "debit": ["制造费用"]},
    {"credit": ["合同结算"], "debit": ["应收账款"]},
    {"credit": ["无形资产"], "debit": ["持有待售资产"]},
    {"credit": ["无形资产减值准备"], "debit": ["资产减值损失"]},
    {"credit": ["应付账款"], "debit": ["原材料"]},
    {"credit": ["债权投资减值准备"], "debit": ["资产减值损失"]},
    {"credit": ["长期借款"], "debit": ["银行存款"]},
    {"credit": ["银行存款"], "debit": ["递延收益"]},
    {"credit": ["预付账款"], "debit": ["材料采购"]},
    {"credit": ["应收账款"], "debit": ["坏账准备"]},
    {"credit": ["银行存款"], "debit": ["财务费用"]},
    {"credit": ["长期股权投资"], "debit": ["银行存款"]},
    {"credit": ["原材料"], "debit": ["研发支出"]},
    {"credit": ["银行存款"], "debit": ["预计负债"]},
    {"credit": ["其他综合收益"], "debit": ["长期股权投资"]},
    {"credit": ["应付职工薪酬"], "debit": ["销售费用"]},
    {"credit": ["递延收益"], "debit": ["银行存款"]},
    {"credit": ["原材料"], "debit": ["生产成本"]},
    {"credit": ["银行存款"], "debit": ["应付账款"]},
    {"credit": ["材料成本差异"], "debit": ["材料采购"]},
    {"credit": ["主营业务收入"], "debit": ["预收账款"]},
    {"credit": ["研发支出"], "debit": ["管理费用"]},
    {"credit": ["银行存款"], "debit": ["债权投资"]},
]

# 预编译关键词列表
UNIQUE_DEBIT_KWS: List[str] = list(set(
    debit_kw for rule in STANDARD_RULES for debit_kw in rule["debit"]
))
UNIQUE_CREDIT_KWS: List[str] = list(set(
    credit_kw for rule in STANDARD_RULES for credit_kw in rule["credit"]
))
RULE_PAIRS: Set[Tuple[str, str]] = set(
    (debit_kw, credit_kw)
    for rule in STANDARD_RULES
    for debit_kw in rule["debit"]
    for credit_kw in rule["credit"]
)

@functools.lru_cache(maxsize=None)
def check_rule_match(debit_subj: str, credit_subj: str) -> bool:
    """
    检查给定的借方科目和贷方科目是否匹配任意一条规则。
    使用缓存和预编译集合优化性能。
    """
    debit_subj = "" if debit_subj is None else str(debit_subj)
    credit_subj = "" if credit_subj is None else str(credit_subj)

    matched_dkws = [kw for kw in UNIQUE_DEBIT_KWS if kw in debit_subj]
    if not matched_dkws:
        return False
        
    matched_ckws = [kw for kw in UNIQUE_CREDIT_KWS if kw in credit_subj]
    if not matched_ckws:
        return False
        
    for d_kw in matched_dkws:
        for c_kw in matched_ckws:
            if (d_kw, c_kw) in RULE_PAIRS:
                return True
                
    return False

def show_column_mapping_dialog(all_columns: List[str], required_columns: List[str], parent=None) -> Optional[Dict[str, str]]:
    """显示列映射对话框，返回映射字典。"""
    if not USE_CUSTOMTKINTER and not tk:
        return None

    mapping = {}

    if USE_CUSTOMTKINTER:
        dialog = ctk.CTkToplevel() if parent else ctk.CTk()
        dialog.title("列映射配置")
        dialog.geometry("550x450")

        lbl_desc = ctk.CTkLabel(dialog, text="检测到输入文件缺少必要列，请手动建立映射关系：",
                                font=("微软雅黑", 12), text_color="#FF6B6B")
        lbl_desc.pack(pady=10, padx=10)

        frame_container = ctk.CTkFrame(dialog)
        frame_container.pack(fill="both", expand=True, padx=10, pady=5)

        canvas = tk.Canvas(frame_container)
        scrollbar = tk.Scrollbar(frame_container, orient="vertical", command=canvas.yview)
        scrollable_frame = ctk.CTkFrame(canvas, fg_color="transparent")

        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        auto_match_keywords = {
            '账套': ['账套', '核算账套', '公司账套', '账套名', '公司', '核算主体', '主体'],
            '会计月': ['月', '日期', 'date', '期间', 'time', '制单日期', '业务日期'],
            '凭证种类': ['种类', '类型', 'type', 'category', '凭证字'],
            '凭证编号': ['编号', '凭证号', 'number', 'no'],
            '一级科目': ['科目', 'subject', '一级', '名称'],
            '借方发生额': ['借方', 'debit', '借', 'jf'],
            '贷方发生额': ['贷方', 'credit', '贷', 'df']
        }

        comboboxes = {}
        for idx, req_col in enumerate(required_columns):
            row_frame = ctk.CTkFrame(scrollable_frame, fg_color="transparent")
            row_frame.pack(fill="x", pady=5)
            display_name = req_col
            if req_col == '会计月':
                display_name = '会计月/日期'
            lbl = ctk.CTkLabel(row_frame, text=f"{display_name} 对应:", width=20, anchor="e", font=("微软雅黑", 11))
            lbl.pack(side="left")
            combo_values = list(all_columns)
            if req_col == '凭证种类':
                combo_values = ["无/不适用"] + combo_values

            cb = ctk.CTkComboBox(row_frame, values=combo_values, width=250, font=("微软雅黑", 11))
            cb.pack(side="left", padx=10)

            matched = False
            if req_col in all_columns:
                cb.set(req_col)
                matched = True
            else:
                keywords = auto_match_keywords.get(req_col, [])
                for col in all_columns:
                    col_lower = str(col).lower()
                    for kw in keywords:
                        if kw in col_lower:
                            cb.set(col)
                            matched = True
                            break
                    if matched: break
            comboboxes[req_col] = cb

        result = [None]
        def on_confirm():
            current_mapping = {}
            missing_selections = []
            for req_col, cb in comboboxes.items():
                val = cb.get()
                if not val:
                    missing_selections.append(req_col)
                else:
                    current_mapping[req_col] = val
            if missing_selections:
                CustomMessageBox.showwarning("提示", f"请为以下列选择映射：\n{', '.join(missing_selections)}")
                return
            result[0] = current_mapping
            dialog.destroy()

        def on_cancel():
            dialog.destroy()

        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_frame.pack(pady=10)
        ctk.CTkButton(btn_frame, text="确定", width=80, command=on_confirm, font=("微软雅黑", 11)).pack(side="left", padx=20)
        ctk.CTkButton(btn_frame, text="取消", width=80, command=on_cancel, font=("微软雅黑", 11)).pack(side="left", padx=20)
        dialog.wait_window()
        return result[0]
    else:
        dialog = tk.Toplevel(parent) if parent else tk.Tk()
        dialog.title("列映射配置")
        dialog.geometry("500x400")
        dialog.grab_set()

        lbl_desc = tk.Label(dialog, text="检测到输入文件缺少必要列，请手动建立映射关系：",
                            wraplength=480, justify="left", fg="red", font=("微软雅黑", 11))
        lbl_desc.pack(pady=10, padx=10)

        frame_container = tk.Frame(dialog)
        frame_container.pack(fill="both", expand=True, padx=10, pady=5)

        canvas = tk.Canvas(frame_container)
        scrollbar = tk.Scrollbar(frame_container, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)

        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        auto_match_keywords = {
            '账套': ['账套', '核算账套', '公司账套', '账套名', '公司', '核算主体', '主体'],
            '会计月': ['月', '日期', 'date', '期间', 'time', '制单日期', '业务日期'],
            '凭证种类': ['种类', '类型', 'type', 'category', '凭证字'],
            '凭证编号': ['编号', '凭证号', 'number', 'no'],
            '一级科目': ['科目', 'subject', '一级', '名称'],
            '借方发生额': ['借方', 'debit', '借', 'jf'],
            '贷方发生额': ['贷方', 'credit', '贷', 'df']
        }

        comboboxes = {}
        for idx, req_col in enumerate(required_columns):
            row_frame = tk.Frame(scrollable_frame)
            row_frame.pack(fill="x", pady=5)
            display_name = req_col
            if req_col == '会计月':
                display_name = '会计月/日期'
            lbl = tk.Label(row_frame, text=f"{display_name} 对应:", width=15, anchor="e", font=("微软雅黑", 10))
            lbl.pack(side="left")
            combo_values = list(all_columns)
            if req_col == '凭证种类':
                combo_values = ["无/不适用"] + combo_values

            cb = ttk.Combobox(row_frame, values=combo_values, state="readonly", width=35, font=("微软雅黑", 10))
            cb.pack(side="left", padx=10)

            matched = False
            if req_col in all_columns:
                cb.set(req_col)
                matched = True
            else:
                keywords = auto_match_keywords.get(req_col, [])
                for col in all_columns:
                    col_lower = str(col).lower()
                    for kw in keywords:
                        if kw in col_lower:
                            cb.set(col)
                            matched = True
                            break
                    if matched: break
            comboboxes[req_col] = cb

        result = [None]
        def on_confirm():
            current_mapping = {}
            missing_selections = []
            for req_col, cb in comboboxes.items():
                val = cb.get()
                if not val:
                    missing_selections.append(req_col)
                else:
                    current_mapping[req_col] = val
            if missing_selections:
                messagebox.showwarning("提示", f"请为以下列选择映射：\n{', '.join(missing_selections)}")
                return
            result[0] = current_mapping
            dialog.destroy()

        def on_cancel():
            dialog.destroy()

        btn_frame = tk.Frame(dialog)
        btn_frame.pack(pady=10)
        tk.Button(btn_frame, text="确定", command=on_confirm, width=10, font=("微软雅黑", 10)).pack(side="left", padx=20)
        tk.Button(btn_frame, text="取消", command=on_cancel, width=10, font=("微软雅黑", 10)).pack(side="left", padx=20)
        dialog.wait_window()
        return result[0]

# ==========================================
# 核心处理逻辑
# ==========================================

def get_leading_digit(n: Any) -> Optional[int]:
    """提取数字的首位数（用于班福定律分析）。"""
    try:
        if pd.isna(n): return None
        if isinstance(n, (int, float)):
            if n == 0: return None
            s = f"{abs(n):.15g}".lower() # 使用科学计数法格式化，避免浮点精度问题
        else:
            s = str(n).replace(',', '').strip().lower()
            
        if 'e' in s:
            base = s.split('e')[0]
            base = base.replace('.', '').replace('-', '')
            for char in base:
                if char.isdigit() and char != '0':
                    return int(char)
        else:
            s = s.replace('.', '').replace('-', '')
            for char in s:
                if char.isdigit() and char != '0':
                    return int(char)
        return None
    except (ValueError, TypeError, AttributeError):
        return None


# ==========================================
# 折半枚举算法（Meet-in-the-Middle）
# ==========================================

def _generate_left_combos(
    candidates: List[Tuple[Any, int]], 
    target_li: int, 
    tolerance_li: int,
    max_iterations: int = MITM_ITERATION_LIMIT
) -> Tuple[List[Tuple[int, List[Any]]], int]:
    """
    生成左半区所有合法组合。
    
    :param candidates: 左半区候选列表 [(对象, 金额厘), ...]
    :param target_li: 目标金额（厘）
    :param tolerance_li: 容差（厘）
    :param max_iterations: 最大迭代次数
    :return: (组合列表, 迭代次数) - 组合列表按金额排序
    """
    n = len(candidates)
    if n == 0:
        return [(0, [])], 0
    
    combos: List[Tuple[int, List[Any]]] = []
    iterations = 0
    upper_bound = target_li + tolerance_li
    
    def dfs(idx: int, current_sum: int, path: List[Any]):
        nonlocal iterations
        iterations += 1
        if iterations > max_iterations:
            return
        
        if current_sum > upper_bound:
            return
        
        if idx == n:
            combos.append((current_sum, path.copy()))
            return
        
        obj, amt_li = candidates[idx]
        
        path.append(obj)
        dfs(idx + 1, current_sum + amt_li, path)
        path.pop()
        
        dfs(idx + 1, current_sum, path)
    
    dfs(0, 0, [])
    combos.sort(key=lambda x: x[0])
    return combos, iterations


def solve_subset_sum_mitm(
    target: float, 
    candidates: List[Tuple[Any, Dict[str, Any]]], 
    tolerance: float = 0.01
) -> Optional[List[Any]]:
    """
    使用折半枚举（Meet-in-the-Middle）算法求解子集和问题。
    
    时间复杂度：O(2^(N/2) * log(2^(N/2)))，相比回溯法O(2^N)大幅优化。
    适用于候选数量较多的"长尾发票凑单"场景。
    
    :param target: 目标金额（元）
    :param candidates: 候选列表，每个元素为 (item_obj, {'amount': 金额})
    :param tolerance: 容差（元），默认0.01元=1分
    :return: 匹配的 item_obj 列表，如果无解返回 None
    """
    target_li = PrecisionEngine.to_integer_li(target)
    tolerance_li = PrecisionEngine.to_integer_li(tolerance)
    
    if not candidates:
        return None
    
    valid_candidates = []
    for cand in candidates:
        if len(cand) < 2 or not isinstance(cand[1], dict):
            continue
        amt = cand[1].get('amount', 0)
        amt_li = PrecisionEngine.to_integer_li(amt)
        if amt_li <= target_li + tolerance_li and amt_li > 0:
            valid_candidates.append((cand, amt_li))
    
    if not valid_candidates:
        return None
    
    valid_candidates.sort(key=lambda x: x[1], reverse=True)
    
    total_li = sum(x[1] for x in valid_candidates)
    if total_li < target_li - tolerance_li:
        return None
    
    if len(valid_candidates) > MAX_CANDIDATE_POOL_SIZE:
        valid_candidates = valid_candidates[:MAX_CANDIDATE_POOL_SIZE]
    
    n = len(valid_candidates)
    mid = n // 2
    
    left_raw = valid_candidates[:mid]
    right_raw = valid_candidates[mid:]
    
    left_combos, left_iters = _generate_left_combos(
        [(c[0], c[1]) for c in left_raw], 
        target_li, 
        tolerance_li
    )
    
    if not left_combos:
        return None
    
    left_values = [c[0] for c in left_combos]
    
    right_iterations = 0
    
    def search_right(idx: int, current_sum: int, path: List[Any]) -> Optional[List[Any]]:
        nonlocal right_iterations
        right_iterations += 1
        if right_iterations > MITM_ITERATION_LIMIT:
            return None
        
        if current_sum > target_li + tolerance_li:
            return None
        
        lower_bound = target_li - current_sum - tolerance_li
        upper_bound = target_li - current_sum + tolerance_li
        
        left_idx = bisect.bisect_left(left_values, lower_bound)
        right_idx = bisect.bisect_right(left_values, upper_bound)
        
        if left_idx < right_idx:
            return left_combos[left_idx][1] + path
        
        if idx == len(right_raw):
            return None
        
        obj, amt_li = right_raw[idx]
        
        path.append(obj)
        result = search_right(idx + 1, current_sum + amt_li, path)
        if result is not None:
            return result
        path.pop()
        
        return search_right(idx + 1, current_sum, path)
    
    result = search_right(0, 0, [])
    return result


def solve_subset_sum(target: float, candidates: List[Tuple[Any, float]], tolerance: float = TOLERANCE) -> Optional[List[Any]]:
    """
    求解子集和问题（使用折半枚举算法）。
    
    :param target: 目标金额
    :param candidates: 候选列表，每个元素为 (item_obj, amount_dict)
    :param tolerance: 容差（元）
    :return: 匹配的 item_obj 列表，如果无解返回 None
    """
    return solve_subset_sum_mitm(target, candidates, tolerance)

class GroupProcessor:
    """
    封装凭证分组处理逻辑，提高可读性和维护性。
    """
    def __init__(self, group_data: Tuple[Any, pd.DataFrame]):
        self.name, self.group = group_data
        self.output_rows: List[Dict[str, Any]] = []
        self.debit_items: List[Dict[str, Any]] = []
        self.credit_items: List[Dict[str, Any]] = []
        self.all_items: List[Dict[str, Any]] = []
        
    def process(self) -> List[Dict[str, Any]]:
        # 0. 预处理：转换数据
        self._prepare_data()
        
        # 1. 阶段零：结转损益特殊规则
        if self._process_profit_loss():
            return self.output_rows
            
        # 2. 阶段一：标准模板优先匹配
        self._match_standard_rules()
        
        # 3. 阶段二：算法金额匹配
        self._match_algo()
        
        # 4. 阶段三：兜底处理
        self._match_fallback()
        
        return self.output_rows

    def _prepare_data(self):
        records = self.group.to_dict('records')
        for row in records:
            d = row['借方发生额']
            c = row['贷方发生额']
            subject = str(row['一级科目'])
            item = {
                'row_data': row.copy(),
                'subject': subject,
                'matched': False
            }
            if d > 0:
                item['amount'] = float(PrecisionEngine.to_decimal(d))
                item['amount_li'] = PrecisionEngine.to_integer_li(d)
                item['side'] = 'debit'
                self.debit_items.append(item)
            elif c < 0:
                item['amount'] = float(PrecisionEngine.to_decimal(abs(c)))
                item['amount_li'] = PrecisionEngine.to_integer_li(abs(c))
                item['side'] = 'debit' 
                self.debit_items.append(item)
            elif c > 0:
                item['amount'] = float(PrecisionEngine.to_decimal(c))
                item['amount_li'] = PrecisionEngine.to_integer_li(c)
                item['side'] = 'credit'
                self.credit_items.append(item)
            elif d < 0:
                item['amount'] = float(PrecisionEngine.to_decimal(abs(d)))
                item['amount_li'] = PrecisionEngine.to_integer_li(abs(d))
                item['side'] = 'credit' 
                self.credit_items.append(item)
            else:
                item['amount'] = 0
                item['amount_li'] = 0
                item['side'] = 'skip'
                row_out = row.copy()
                row_out['对方科目'] = None
                row_out['匹配类型'] = '无发生额'
                self.output_rows.append(row_out)
        
        self.debit_items.sort(key=lambda x: x['amount_li'], reverse=True)
        self.credit_items.sort(key=lambda x: x['amount_li'], reverse=True)
        self.all_items = self.debit_items + self.credit_items

    def _add_match(self, item1: Dict, item2_list: List[Dict], match_type: str = "算法生成"):
        for match_item in item2_list:
            row_out = item1['row_data'].copy()
            row_out['对方科目'] = match_item['subject']
            row_out['匹配类型'] = match_type
            amt = match_item['amount']
            if item1['side'] == 'debit':
                if row_out['借方发生额'] > 0: row_out['借方发生额'] = amt
                else: row_out['贷方发生额'] = -amt
            else:
                if row_out['贷方发生额'] > 0: row_out['贷方发生额'] = amt
                else: row_out['借方发生额'] = -amt
            self.output_rows.append(row_out)
            
            row_out_2 = match_item['row_data'].copy()
            row_out_2['对方科目'] = item1['subject']
            row_out_2['匹配类型'] = match_type
            self.output_rows.append(row_out_2)

    def _is_profit_subject(self, subj: str) -> bool:
        s = str(subj).strip()
        if any(x in s for x in ["费用", "成本", "折旧", "减值", "摊销"]):
            return False
        return s.startswith("本年利润") or s.startswith("利润分配")

    def _process_profit_loss(self) -> bool:
        profit_items = [x for x in self.all_items if self._is_profit_subject(x['subject'])]
        if not profit_items:
            return False
        if len(profit_items) > 1:
            # 多条损益科目时，特殊规则会错误吞并到第一条；回退到通用匹配更安全。
            return False
            
        other_items = [x for x in self.all_items if x not in profit_items]
        if not other_items:
            return False

        for item in self.all_items: item['matched'] = True
        
        p_debit_sum = sum(p['amount'] for p in profit_items if p['side'] == 'debit')
        p_credit_sum = sum(p['amount'] for p in profit_items if p['side'] == 'credit')
        p_net = p_debit_sum - p_credit_sum
        
        target_col = '借方发生额' if p_net >= -FLOAT_EPSILON else '贷方发生额'
        calc_p_net = 0

        for o_item in other_items:
            row_o = o_item['row_data'].copy()
            row_o['对方科目'] = profit_items[0]['subject']
            row_o['匹配类型'] = '结转损益规则'
            self.output_rows.append(row_o)
            
            row_p = profit_items[0]['row_data'].copy()
            row_p['对方科目'] = o_item['subject']
            row_p['匹配类型'] = '结转损益规则(拆分)'
            row_p['借方发生额'] = 0
            row_p['贷方发生额'] = 0
            
            val = o_item['amount']
            is_o_debit = (o_item['side'] == 'debit')
            should_be_debit = not is_o_debit
            
            if should_be_debit: calc_p_net += val
            else: calc_p_net -= val
            
            is_target_debit = (target_col == '借方发生额')
            final_val = 0
            if is_target_debit:
                final_val = val if should_be_debit else -val
            else:
                final_val = -val if should_be_debit else val
            row_p[target_col] = final_val
            self.output_rows.append(row_p)

        diff = p_net - calc_p_net
        if not PrecisionEngine.amounts_match(diff, 0, tolerance_li=100):
            row_diff = profit_items[0]['row_data'].copy()
            row_diff['对方科目'] = '损益结转误差调整'
            row_diff['匹配类型'] = '结转损益规则(调整)'
            row_diff['借方发生额'] = 0
            row_diff['贷方发生额'] = 0
            if target_col == '借方发生额': row_diff['借方发生额'] = diff
            else: row_diff['贷方发生额'] = -diff
            self.output_rows.append(row_diff)
        return True

    def _match_standard_rules(self):
        # 1.1 一对一
        for d in self.debit_items:
            if d['matched']: continue
            for c in self.credit_items:
                if c['matched']: continue
                if PrecisionEngine.amounts_match(d['amount'], c['amount']) and check_rule_match(d['subject'], c['subject']):
                    d['matched'] = True
                    c['matched'] = True
                    self._add_match(d, [c], match_type="标准模板")
                    break
        # 1.2 一对多
        for d in self.debit_items:
            if d['matched']: continue
            candidates = [(i, c) for i, c in enumerate(self.credit_items) 
                          if not c['matched'] and PrecisionEngine.compare_amounts(c['amount'], d['amount']) <= 0
                          and check_rule_match(d['subject'], c['subject'])]
            match = solve_subset_sum(d['amount'], candidates)
            if match:
                d['matched'] = True
                matched_credits = []
                for idx, c in match:
                    c['matched'] = True
                    matched_credits.append(c)
                self._add_match(d, matched_credits, match_type="标准模板")
        # 1.3 多对一
        for c in self.credit_items:
            if c['matched']: continue
            candidates = [(i, d) for i, d in enumerate(self.debit_items) 
                          if not d['matched'] and PrecisionEngine.compare_amounts(d['amount'], c['amount']) <= 0
                          and check_rule_match(d['subject'], c['subject'])]
            match = solve_subset_sum(c['amount'], candidates)
            if match:
                c['matched'] = True
                matched_debits = []
                for idx, d in match:
                    d['matched'] = True
                    matched_debits.append(d)
                self._add_match(c, matched_debits, match_type="标准模板")

    def _match_algo(self):
        # 2.1 一对一
        for d in self.debit_items:
            if d['matched']: continue
            for c in self.credit_items:
                if c['matched']: continue
                if PrecisionEngine.amounts_match(d['amount'], c['amount']):
                    d['matched'] = True
                    c['matched'] = True
                    self._add_match(d, [c], match_type="算法生成")
                    break
        # 2.2 一对多
        for d in self.debit_items:
            if d['matched']: continue
            candidates = [(i, c) for i, c in enumerate(self.credit_items) 
                          if not c['matched'] and PrecisionEngine.compare_amounts(c['amount'], d['amount']) <= 0]
            match = solve_subset_sum(d['amount'], candidates)
            if match:
                d['matched'] = True
                matched_credits = []
                for idx, c in match:
                    c['matched'] = True
                    matched_credits.append(c)
                self._add_match(d, matched_credits, match_type="算法生成")
        # 2.3 多对一
        for c in self.credit_items:
            if c['matched']: continue
            candidates = [(i, d) for i, d in enumerate(self.debit_items) 
                          if not d['matched'] and PrecisionEngine.compare_amounts(d['amount'], c['amount']) <= 0]
            match = solve_subset_sum(c['amount'], candidates)
            if match:
                c['matched'] = True
                matched_debits = []
                for idx, d in match:
                    d['matched'] = True
                    matched_debits.append(d)
                self._add_match(c, matched_debits, match_type="算法生成")

    def _match_fallback(self):
        unmatched_debit = [d for d in self.debit_items if not d['matched']]
        unmatched_credit = [c for c in self.credit_items if not c['matched']]
        d_idx = 0
        c_idx = 0
        
        while d_idx < len(unmatched_debit) and c_idx < len(unmatched_credit):
            d_item = unmatched_debit[d_idx]
            c_item = unmatched_credit[c_idx]
            if 'rem_amount' not in d_item: d_item['rem_amount'] = d_item['amount']
            if 'rem_amount' not in c_item: c_item['rem_amount'] = c_item['amount']
            val_d = d_item['rem_amount']
            val_c = c_item['rem_amount']
            matched_val = min(val_d, val_c)
            
            def set_split_amount(row_data, val):
                d_orig = row_data.get('借方发生额', 0)
                c_orig = row_data.get('贷方发生额', 0)
                if abs(d_orig) > FLOAT_EPSILON:
                    sign = 1 if d_orig > 0 else -1
                    row_data['借方发生额'] = val * sign
                elif abs(c_orig) > FLOAT_EPSILON:
                    sign = 1 if c_orig > 0 else -1
                    row_data['贷方发生额'] = val * sign

            row_d = d_item['row_data'].copy()
            row_d['对方科目'] = c_item['subject']
            row_d['匹配类型'] = '算法生成(兜底拆分)'
            set_split_amount(row_d, matched_val)
            self.output_rows.append(row_d)
            
            row_c = c_item['row_data'].copy()
            row_c['对方科目'] = d_item['subject']
            row_c['匹配类型'] = '算法生成(兜底拆分)'
            set_split_amount(row_c, matched_val)
            self.output_rows.append(row_c)
            
            if PrecisionEngine.amounts_match(val_d, matched_val): d_idx += 1
            else: d_item['rem_amount'] -= matched_val
            if PrecisionEngine.amounts_match(val_c, matched_val): c_idx += 1
            else: c_item['rem_amount'] -= matched_val
        
        for i in range(d_idx, len(unmatched_debit)):
            d = unmatched_debit[i]
            row = d['row_data'].copy()
            row['对方科目'] = '未找到匹配'
            row['匹配类型'] = '算法生成(异常剩余)'
            rem = d.get('rem_amount', d['amount'])
            d_orig = row.get('借方发生额', 0)
            c_orig = row.get('贷方发生额', 0)
            if abs(d_orig) > FLOAT_EPSILON:
                sign = 1 if d_orig > 0 else -1
                row['借方发生额'] = rem * sign
            elif abs(c_orig) > FLOAT_EPSILON:
                sign = 1 if c_orig > 0 else -1
                row['贷方发生额'] = rem * sign
            self.output_rows.append(row)
            
        for i in range(c_idx, len(unmatched_credit)):
            c = unmatched_credit[i]
            row = c['row_data'].copy()
            row['对方科目'] = '未找到匹配'
            row['匹配类型'] = '算法生成(异常剩余)'
            rem = c.get('rem_amount', c['amount'])
            d_orig = row.get('借方发生额', 0)
            c_orig = row.get('贷方发生额', 0)
            if abs(d_orig) > FLOAT_EPSILON:
                sign = 1 if d_orig > 0 else -1
                row['借方发生额'] = rem * sign
            elif abs(c_orig) > FLOAT_EPSILON:
                sign = 1 if c_orig > 0 else -1
                row['贷方发生额'] = rem * sign
            self.output_rows.append(row)

def process_group(group_data: Tuple[Any, pd.DataFrame]) -> List[Dict[str, Any]]:
    processor = GroupProcessor(group_data)
    return processor.process()

# ==========================================
# 主流程模块化重构
# ==========================================

def detect_ledger_account_column(columns: List[str]) -> Optional[str]:
    """
    检测账套列（支持多种列名变体）。
    
    :param columns: 文件列名列表
    :return: 检测到的账套列名，如果未找到返回 None
    """
    columns_lower = {str(col).lower().strip(): col for col in columns}
    for keyword in LEDGER_ACCOUNT_KEYWORDS:
        keyword_lower = keyword.lower()
        for col_lower, col_original in columns_lower.items():
            if keyword_lower in col_lower:
                return col_original
            if col_lower in keyword_lower:
                return col_original
        if keyword in columns:
            return keyword
    return None

def check_column_mapping_needed(input_path: str) -> Tuple[bool, List[str], List[str]]:
    """检查是否需要列映射，返回(是否需要映射, 文件列名, 缺少的列名)"""
    required_columns = ['会计月', '凭证种类', '凭证编号', '一级科目', '借方发生额', '贷方发生额']
    try:
        try:
            import python_calamine
            engine = 'calamine'
        except ImportError:
            engine = 'openpyxl'
        df = pd.read_excel(input_path, engine=engine, nrows=0)
        file_columns = list(df.columns)
        missing_cols = [col for col in required_columns if col not in file_columns]
        return len(missing_cols) > 0, file_columns, missing_cols
    except Exception as e:
        print(f"检查列名失败: {e}")
        return False, [], []

def apply_column_mapping(df: pd.DataFrame, mapping: Dict[str, str]) -> pd.DataFrame:
    """应用列映射到DataFrame"""
    valid_mapping = {k: v for k, v in mapping.items() if v != "无/不适用"}
    df.rename(columns={v: k for k, v in valid_mapping.items()}, inplace=True)
    
    if mapping.get('凭证种类') == "无/不适用" and '凭证种类' in df.columns:
        df.drop(columns=['凭证种类'], inplace=True)
    
    return df

def load_and_preprocess_data(input_path: str, interactive: bool) -> Optional[pd.DataFrame]:
    """加载并预处理数据。"""
    GUI_PROGRESS.update(0, "初始化...", "准备阶段")
    print(f"正在读取文件: {input_path}")
    GUI_PROGRESS.update(5, "正在读取Excel文件...", "数据读取")
    
    try:
        try:
            import python_calamine
            engine = 'calamine'
        except ImportError:
            engine = 'openpyxl'
        print(f"使用读取引擎: {engine}")
        df = pd.read_excel(input_path, engine=engine)
        GUI_PROGRESS.update(15, "文件读取完成", "数据读取")
    except Exception as e:
        print(f"读取Excel文件失败: {e}")
        GUI_PROGRESS.stop()
        return None

    required_columns = ['会计月', '凭证种类', '凭证编号', '一级科目', '借方发生额', '贷方发生额']
    missing_cols = [col for col in required_columns if col not in df.columns]

    if missing_cols:
        print(f"输入文件缺少必要列: {missing_cols}")
        if interactive and (USE_CUSTOMTKINTER or tk):
            GUI_PROGRESS.update(16, "等待用户配置列映射...", "数据预处理")
            print("等待用户手动配置列映射...")
            mapping = show_column_mapping_dialog(list(df.columns), required_columns)
            if mapping:
                print(f"用户已配置列映射: {mapping}")
                valid_mapping = {k: v for k, v in mapping.items() if v != "无/不适用"}
                df.rename(columns={v: k for k, v in valid_mapping.items()}, inplace=True)
                
                if mapping.get('凭证种类') == "无/不适用" and '凭证种类' in df.columns:
                     print("用户选择忽略凭证种类，正在移除该列...")
                     df.drop(columns=['凭证种类'], inplace=True)
                
                missing_cols = [col for col in required_columns if col not in df.columns]
                if '凭证种类' in missing_cols and mapping.get('凭证种类') == "无/不适用":
                    missing_cols.remove('凭证种类')
                
                if missing_cols:
                    print(f"错误: 即使映射后仍缺少列: {missing_cols}")
                    GUI_PROGRESS.stop()
                    return None
            else:
                print("用户取消了列映射配置。")
                GUI_PROGRESS.stop()
                return None
        else:
            print("错误: 缺少必要列且未启用交互模式或 GUI 不可用。")
            GUI_PROGRESS.stop()
            return None

    # 数据填充与清洗
    cols_to_ffill = ['会计月', '凭证种类', '凭证编号']
    
    # 检测账套列并加入填充列表
    ledger_col = detect_ledger_account_column(list(df.columns))
    if ledger_col:
        print(f"检测到账套列: [{ledger_col}]，将作为分组条件之一")
        cols_to_ffill.insert(0, ledger_col)
        df.attrs['ledger_column'] = ledger_col
    else:
        df.attrs['ledger_column'] = None
    
    for col in cols_to_ffill:
        if col in df.columns:
            df[col] = df[col].replace(r'^\s*$', float('nan'), regex=True)
            if df[col].isnull().any():
                print(f"检测到 [{col}] 列存在空值，正在尝试向下填充(处理合并单元格)...")
                df[col] = df[col].ffill()

    if '会计月' in df.columns:
        print("注意：已禁用日期格式自动转换，将直接使用原文件中的[会计月]文本进行分组。")
        df['会计月'] = df['会计月'].astype(str).str.strip()

    df['借方发生额'] = pd.to_numeric(df['借方发生额'], errors='coerce').fillna(0)
    df['贷方发生额'] = pd.to_numeric(df['贷方发生额'], errors='coerce').fillna(0)
    
    print("数据预处理完成。")
    return df

def get_anomaly_threshold(interactive: bool) -> float:
    """获取异常分录筛选阈值。"""
    anomaly_threshold = 10000.0
    if interactive and (USE_CUSTOMTKINTER or tk):
        root = None
        try:
            if USE_CUSTOMTKINTER:
                root = ctk.CTk()
                root.withdraw()
                dialog = CustomInputDialog(root, "配置", "请输入异常分录筛选阈值 (金额):", initial_value=10000.0, min_value=0.0)
                user_val = dialog.get_result()
                if user_val is not None: anomaly_threshold = user_val
            else:
                root = tk.Tk()
                root.withdraw()
                user_val = simpledialog.askfloat("配置", "请输入异常分录筛选阈值 (金额):", initialvalue=10000.0, minvalue=0.0)
                if user_val is not None: anomaly_threshold = user_val
        finally:
            if root: root.destroy()
    print(f"异常分录筛选阈值设定为: {anomaly_threshold}")
    return anomaly_threshold

def perform_processing(df: pd.DataFrame) -> Tuple[pd.DataFrame, List[Dict]]:
    """执行并行计算处理。"""
    group_cols = ['会计月', '凭证编号']
    type_stats = ""
    if '凭证种类' in df.columns:
        group_cols.insert(1, '凭证种类')
        type_stats = f", 凭证种类({df['凭证种类'].nunique()}个)"
    
    # 检测并添加账套列到分组条件
    ledger_col = df.attrs.get('ledger_column', None)
    if ledger_col and ledger_col in df.columns:
        group_cols.insert(0, ledger_col)
        type_stats = f", 账套({df[ledger_col].nunique()}个)" + type_stats
        
    print(f"分组字段唯一值统计: 会计月({df['会计月'].nunique()}个){type_stats}, 凭证编号({df['凭证编号'].nunique()}个)")
    grouped = df.groupby(group_cols)
    total_groups = len(grouped)
    
    if df['会计月'].nunique() < 2 and total_groups > 10:
        print("警告：检测到【会计月】列只有一个唯一值，如果您的数据包含多个月份，说明月份列可能被错误修改或覆盖！")
    
    tasks = [(name, group) for name, group in grouped]
    output_rows = []
    failed_groups = []
    processed_count = 0
    start_pct = 15
    end_pct = 90
    max_workers = max(1, (os.cpu_count() or 4) - 1)
    print(f"启动 {max_workers} 个进程进行并行计算...")
    
    with ProcessPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(process_group, task): task for task in tasks}
        for future in as_completed(futures):
            try:
                result = future.result()
                output_rows.extend(result)
            except Exception as e:
                task_name = futures[future][0]
                print(f"处理分组 {task_name} 失败: {e}")
                traceback.print_exc()
                failed_groups.append({
                    "分组键": str(task_name),
                    "错误": str(e),
                    "堆栈": traceback.format_exc()
                })
            
            processed_count += 1
            if processed_count % 100 == 0 or processed_count == total_groups:
                current_pct = start_pct + (processed_count / total_groups) * (end_pct - start_pct)
                GUI_PROGRESS.update(current_pct, f"已处理 {processed_count}/{total_groups}", "并行计算中")

    print(f"生成输出数据，共 {len(output_rows)} 行...")
    out_df = pd.DataFrame(output_rows)
    return out_df, failed_groups

def validate_results(df: pd.DataFrame, out_df: pd.DataFrame):
    """验证结果的完整性。"""
    print("正在进行数据完整性校验...")
    GUI_PROGRESS.update(95, "正在校验数据...", "数据校验")
    
    if out_df is None:
        out_df = pd.DataFrame()
    for col in REQUIRED_OUTPUT_COLUMNS:
        if col not in out_df.columns:
            out_df[col] = pd.Series(dtype='object' if col == '对方科目' else 'float64')
    
    orig_debit_sum = df['借方发生额'].sum()
    orig_credit_sum = df['贷方发生额'].sum()
    out_debit_sum = out_df['借方发生额'].sum()
    out_credit_sum = out_df['贷方发生额'].sum()
    diff_debit = abs(orig_debit_sum - out_debit_sum)
    diff_credit = abs(orig_credit_sum - out_credit_sum)
    
    print(f"原借方合计: {orig_debit_sum:,.2f}, 新借方合计: {out_debit_sum:,.2f}, 差额: {diff_debit:,.2f}")
    print(f"原贷方合计: {orig_credit_sum:,.2f}, 新贷方合计: {out_credit_sum:,.2f}, 差额: {diff_credit:,.2f}")
    
    if diff_debit > AMOUNT_CHECK_THRESHOLD or diff_credit > AMOUNT_CHECK_THRESHOLD:
        print("严重警告：金额合计不一致！请检查生成逻辑。")
        GUI_PROGRESS.update(96, "警告：金额校验失败", "校验失败")
    else:
        print("金额校验通过。")
        
    valid_rows = out_df[(abs(out_df['借方发生额']) > AMOUNT_VALID_THRESHOLD) | (abs(out_df['贷方发生额']) > AMOUNT_VALID_THRESHOLD)]
    empty_contra = valid_rows[valid_rows['对方科目'].isnull() | (valid_rows['对方科目'] == '')]
    if not empty_contra.empty:
        print(f"警告：发现 {len(empty_contra)} 行有发生额但无对方科目！")
        GUI_PROGRESS.update(96, "警告：对方科目缺失", "校验失败")
    
    multi_contra = valid_rows[valid_rows['对方科目'].astype(str).str.contains(',')]
    if not multi_contra.empty:
        print(f"警告：发现 {len(multi_contra)} 行包含多个对方科目（含逗号）！")
        GUI_PROGRESS.update(96, "警告：对方科目未拆分", "校验失败")
        
    if empty_contra.empty and multi_contra.empty:
        print("对方科目格式校验通过。")

def detect_anomalies(out_df: pd.DataFrame, threshold: float) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """检测异常分录。"""
    print("正在检测异常分录...")
    anomaly_rows = []
    anomaly_patterns = []

    if not out_df.empty:
        print("正在逐行扫描异常...")
        for row in out_df.itertuples(index=False):
            d_amt = abs(getattr(row, '借方发生额', 0))
            c_amt = abs(getattr(row, '贷方发生额', 0))
            total_amt = d_amt + c_amt
            if total_amt <= threshold: continue
            
            match_type = str(getattr(row, '匹配类型', ''))
            if '标准模板' in match_type or '结转损益' in match_type: continue
            
            subj = str(getattr(row, '一级科目', ''))
            contra = str(getattr(row, '对方科目', ''))
            
            if not contra or contra == 'nan' or contra == 'None':
                anomaly_rows.append(row._asdict())
                continue
                
            debit_subj = ""
            credit_subj = ""
            if d_amt > AMOUNT_VALID_THRESHOLD:
                debit_subj = subj
                credit_subj = contra
            elif c_amt > AMOUNT_VALID_THRESHOLD:
                credit_subj = subj
                debit_subj = contra
            else:
                continue
                
            if check_rule_match(debit_subj, credit_subj): continue
            
            row_with_keys = row._asdict()
            row_with_keys['__debit_key'] = debit_subj
            row_with_keys['__credit_key'] = credit_subj
            anomaly_rows.append(row_with_keys)
            
            desc = str(getattr(row, '摘要', getattr(row, '业务说明', '')))
            try:
                raw_m = getattr(row, '会计月', '')
                m_str = str(raw_m)
                if hasattr(raw_m, 'strftime'): m_str = raw_m.strftime('%Y-%m')
                elif ' ' in m_str: m_str = m_str.split(' ')[0][:7]
                t_str = str(getattr(row, '凭证种类', '')).strip()
                n_str = str(getattr(row, '凭证编号', '')).strip()
                if t_str and t_str != 'nan':
                    full_voucher = f"{m_str}-{t_str}-{n_str}"
                else:
                    full_voucher = f"{m_str}-{n_str}"
            except Exception:
                full_voucher = f"{getattr(row, '会计月', '')}-{getattr(row, '凭证编号', '')}"

            anomaly_patterns.append({
                "借方科目": debit_subj,
                "贷方科目": credit_subj,
                "业务描述": desc,
                "金额": total_amt,
                "完整凭证": full_voucher
            })
            
    if anomaly_rows:
        anomaly_df = pd.DataFrame(anomaly_rows)
        if '会计月' in anomaly_df.columns and '凭证编号' in anomaly_df.columns:
             anomaly_df.sort_values(by=['会计月', '凭证编号'], ascending=[False, False], inplace=True)
        print(f"检测到 {len(anomaly_rows)} 笔异常分录 (金额 > {threshold})")
    else:
        anomaly_df = pd.DataFrame(columns=out_df.columns)
        print("未检测到符合条件的异常分录。")

    aggregated_patterns = pd.DataFrame()
    if anomaly_patterns:
        pat_df = pd.DataFrame(anomaly_patterns)
        pat_df.sort_values(by='金额', ascending=False, kind='mergesort', inplace=True)
        agg_rules = {'金额': 'sum', '业务描述': 'first', '完整凭证': 'first'}
        grouped_pat = pat_df.groupby(['借方科目', '贷方科目'], as_index=False).agg(agg_rules)
        grouped_pat['异常类型编号'] = range(1, len(grouped_pat) + 1)
        grouped_pat.rename(columns={'完整凭证': '凭证编号'}, inplace=True)
        cols_order = ['借方科目', '贷方科目', '业务描述', '金额', '凭证编号', '异常类型编号']
        aggregated_patterns = grouped_pat[cols_order]
    else:
        aggregated_patterns = pd.DataFrame(columns=['借方科目', '贷方科目', '业务描述', '金额', '凭证编号', '异常类型编号'])

    if not anomaly_df.empty and not aggregated_patterns.empty:
        mapping_df = aggregated_patterns[['借方科目', '贷方科目', '异常类型编号']]
        anomaly_df = pd.merge(
            anomaly_df,
            mapping_df,
            left_on=['__debit_key', '__credit_key'],
            right_on=['借方科目', '贷方科目'],
            how='left'
        )
        drop_cols = ['__debit_key', '__credit_key', '借方科目', '贷方科目']
        anomaly_df.drop(columns=[c for c in drop_cols if c in anomaly_df.columns], inplace=True)

    return anomaly_df, aggregated_patterns

def analyze_benford(df: pd.DataFrame) -> pd.DataFrame:
    """班福定律分析。"""
    print("正在计算原始数据的首位数...")
    temp_amount = df['借方发生额'].abs() + df['贷方发生额'].abs()
    df['首位数'] = temp_amount.apply(get_leading_digit)

    stats_data = []
    valid_digits = df['首位数'].dropna()
    total_count = len(valid_digits)
    for d in range(1, 10):
        count = (valid_digits == d).sum()
        actual_pct = count / total_count if total_count > 0 else 0
        benford_pct = math.log10(1 + 1/d)
        stats_data.append({
            "首位数": d,
            "出现次数": count,
            "出现比率": actual_pct,
            "班福比率": benford_pct
        })
    stats_df = pd.DataFrame(stats_data)
    sum_row = pd.DataFrame([{
        "首位数": "合计",
        "出现次数": stats_df["出现次数"].sum(),
        "出现比率": stats_df["出现比率"].sum(),
        "班福比率": stats_df["班福比率"].sum()
    }])
    stats_df = pd.concat([stats_df, sum_row], ignore_index=True)
    return stats_df

def save_output_file(output_path: str, df: pd.DataFrame, out_df: pd.DataFrame, 
                     anomaly_df: pd.DataFrame, aggregated_patterns: pd.DataFrame, 
                     stats_df: pd.DataFrame, failed_groups: List[Dict]):
    """保存结果到 Excel。"""
    print(f"正在保存文件到: {output_path}")
    GUI_PROGRESS.update(92, "正在保存Excel文件...", "文件输出")
    
    if not OPENPYXL_AVAILABLE:
        print("错误: openpyxl库未安装，无法生成Excel文件。请运行: pip install openpyxl")
        GUI_PROGRESS.stop()
        return
    
    try:
        print("正在写入Excel文件(包含多个Sheet及图表)...")
        # 整理列顺序（账套列如果存在则放在最前面）
        ledger_col = df.attrs.get('ledger_column', None)
        cols = []
        if ledger_col and ledger_col in out_df.columns:
            cols.append(ledger_col)
        cols.extend(['记账时间', '会计月', '凭证种类', '凭证编号', '编号', '业务说明', '一级编号', '一级科目', '科目编号', '科目名称', '借方发生额', '贷方发生额', '对方科目', '匹配类型'])
        for c in df.columns:
            if c not in cols and c != '首位数': cols.append(c)
        final_cols = [c for c in cols if c in out_df.columns]
        out_df = out_df[final_cols]

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            out_df.to_excel(writer, sheet_name='生成结果', index=False)
            df.to_excel(writer, sheet_name='原始数据', index=False)
            
            if not anomaly_df.empty:
                anomaly_df.to_excel(writer, sheet_name='异常分录明细', index=False)
            else:
                pd.DataFrame({"提示": ["未检测到符合阈值的异常分录"]}).to_excel(writer, sheet_name='异常分录明细', index=False)
                
            if isinstance(aggregated_patterns, pd.DataFrame) and not aggregated_patterns.empty:
                aggregated_patterns.to_excel(writer, sheet_name='异常分录', index=False)
            else:
                pd.DataFrame(columns=['借方科目', '贷方科目', '业务描述', '金额', '凭证编号', '异常类型编号']).to_excel(writer, sheet_name='异常分录', index=False)

            if failed_groups:
                pd.DataFrame(failed_groups).to_excel(writer, sheet_name='失败分组', index=False)
            
            stats_sheet_name = '班福分析'
            stats_df.to_excel(writer, sheet_name=stats_sheet_name, index=False)
            stats_ws = writer.sheets[stats_sheet_name]

            for row in range(2, 12):
                for col in (3, 4):
                    cell = stats_ws.cell(row=row, column=col)
                    cell.number_format = '0.00%'

            chart = LineChart()
            chart.title = "班福定律分析 - 首位数分布"
            chart.style = 13
            chart.y_axis.title = "比率"
            chart.x_axis.title = "首位数"
            cats = Reference(stats_ws, min_col=1, min_row=2, max_row=10)
            data = Reference(stats_ws, min_col=3, min_row=1, max_row=10, max_col=4)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            stats_ws.add_chart(chart, "F2")
            
        print("完成。")
        GUI_PROGRESS.update(100, "处理完成", "完成")
    except Exception as e:
        print(f"保存文件失败: {e}")
        GUI_PROGRESS.stop()

def run_processing_pipeline(df: pd.DataFrame, anomaly_threshold: float, output_path: str):
    """执行核心处理流水线（适合在工作线程中运行）。"""
    # 3. 执行处理
    pipeline_start = time.time()
    out_df, failed_groups = perform_processing(df)

    # 4. 验证结果
    validate_results(df, out_df)

    # 5. 检测异常
    anomaly_df, aggregated_patterns = detect_anomalies(out_df, anomaly_threshold)

    # 6. 班福分析
    stats_df = analyze_benford(df)

    # 7. 保存结果
    save_output_file(output_path, df, out_df, anomaly_df, aggregated_patterns, stats_df, failed_groups)

    # 8. 输出统计摘要
    pipeline_elapsed = time.time() - pipeline_start
    total_input_rows = len(df)
    total_output_rows = len(out_df)
    matched_count = len(out_df[out_df['匹配类型'] == '标准模板']) if '匹配类型' in out_df.columns else 0
    algo_count = len(out_df[out_df['匹配类型'] == '算法生成']) if '匹配类型' in out_df.columns else 0
    fallback_count = len(out_df[(out_df['匹配类型'].str.contains('兜底拆分|异常剩余', na=False))]) if '匹配类型' in out_df.columns else 0

    print("=" * 50)
    print(f"📊 处理统计摘要")
    print(f"  输入数据行数: {total_input_rows}")
    print(f"  输出数据行数: {total_output_rows}")
    print(f"  标准模板匹配: {matched_count} 行")
    print(f"  算法金额匹配: {algo_count} 行")
    print(f"  兜底/异常处理: {fallback_count} 行")
    print(f"  失败分组数: {len(failed_groups)}")
    print(f"  总耗时: {pipeline_elapsed:.2f} 秒")
    print("=" * 50)

def generate_contra_account(input_path: str, output_path: str, interactive: bool = False):
    """主程序入口 (CLI模式)。"""
    # 1. 加载数据
    df = load_and_preprocess_data(input_path, interactive)
    if df is None: return

    # 2. 配置参数
    anomaly_threshold = get_anomaly_threshold(interactive)

    # 3. 执行流水线
    run_processing_pipeline(df, anomaly_threshold, output_path)

# ==========================================
# GUI 应用入口
# ==========================================

GUI_LOG_QUEUE = queue.Queue()

class GuiLogRedirector:
    """重定向print输出到GUI日志队列"""
    def __init__(self, queue_obj):
        self.queue = queue_obj
    
    def write(self, message):
        if message and message.strip():
            timestamp = datetime.datetime.now().strftime("%H:%M:%S")
            self.queue.put(f"[{timestamp}] {message.strip()}")
    
    def flush(self):
        pass

def run_gui():
    """运行 GUI 应用程序。"""
    global progress_bar
    
    log_redirector = GuiLogRedirector(GUI_LOG_QUEUE)
    sys.stdout = log_redirector
    sys.stderr = log_redirector
    
    if USE_CUSTOMTKINTER:
        ctk.set_appearance_mode("system")
        ctk.set_default_color_theme("blue")
        app = ctk.CTk()
        app.title("对方科目生成工具 v1.8.0")
        app.geometry("640x550")
        app.resizable(True, True)
    else:
        app = tk.Tk()
        app.title("对方科目生成工具 v1.7.0")
        app.geometry("600x520")
        app.resizable(True, True)

    if USE_CUSTOMTKINTER:
        main_frame = ctk.CTkFrame(app)
    else:
        main_frame = tk.Frame(app)
    main_frame.pack(fill="both", expand=True, padx=10, pady=8)

    mapping_container = None
    mapping_comboboxes = {}
    current_input_path = [None]
    current_output_path = [None]
    current_df = [None]
    confirm_btn_widget = [None]

    auto_match_keywords = {
        '账套': ['账套', '核算账套', '公司账套', '账套名', '公司', '核算主体', '主体'],
        '会计月': ['月', '日期', 'date', '期间', 'time', '制单日期', '业务日期'],
        '凭证种类': ['种类', '类型', 'type', 'category', '凭证字'],
        '凭证编号': ['编号', '凭证号', 'number', 'no'],
        '一级科目': ['科目', 'subject', '一级', '名称'],
        '借方发生额': ['借方', 'debit', '借', 'jf'],
        '贷方发生额': ['贷方', 'credit', '贷', 'df']
    }

    def show_mapping_config(file_columns: List[str], missing_cols: List[str]):
        nonlocal mapping_container, mapping_comboboxes
        
        if mapping_container:
            mapping_container.destroy()
        
        if USE_CUSTOMTKINTER:
            mapping_container = ctk.CTkFrame(main_frame)
        else:
            mapping_container = tk.Frame(main_frame)
        mapping_container.pack(fill="x", padx=3, pady=3)
        
        if USE_CUSTOMTKINTER:
            ctk.CTkLabel(mapping_container, text="⚠️ 缺少必要列，请配置映射：", font=("微软雅黑", 10), text_color="#FF6B6B").grid(row=0, column=0, columnspan=3, pady=(3, 2), sticky="w")
        else:
            tk.Label(mapping_container, text="⚠️ 缺少必要列，请配置映射：", font=("微软雅黑", 10), fg="red").grid(row=0, column=0, columnspan=3, pady=(3, 2), sticky="w")
        
        mapping_comboboxes.clear()
        
        for idx, req_col in enumerate(missing_cols):
            row = (idx // 3) + 1
            col = idx % 3
            
            display_name = req_col
            
            if USE_CUSTOMTKINTER:
                lbl = ctk.CTkLabel(mapping_container, text=f"{display_name}:", width=10, anchor="e", font=("微软雅黑", 9))
                lbl.grid(row=row, column=col*2, padx=(5, 0), pady=1, sticky="e")
            else:
                lbl = tk.Label(mapping_container, text=f"{display_name}:", width=8, anchor="e", font=("微软雅黑", 9))
                lbl.grid(row=row, column=col*2, padx=(5, 0), pady=1, sticky="e")
            
            combo_values = list(file_columns)
            if req_col == '凭证种类':
                combo_values = ["无/不适用"] + combo_values
            
            if USE_CUSTOMTKINTER:
                cb = ctk.CTkComboBox(mapping_container, values=combo_values, width=120, font=("微软雅黑", 9))
                cb.grid(row=row, column=col*2+1, padx=(0, 5), pady=1, sticky="w")
            else:
                cb = ttk.Combobox(mapping_container, values=combo_values, state="readonly", width=18, font=("微软雅黑", 9))
                cb.grid(row=row, column=col*2+1, padx=(0, 5), pady=1, sticky="w")
            
            matched = False
            if req_col in file_columns:
                cb.set(req_col)
                matched = True
            else:
                keywords = auto_match_keywords.get(req_col, [])
                for col in file_columns:
                    col_lower = str(col).lower()
                    for kw in keywords:
                        if kw in col_lower:
                            cb.set(col)
                            matched = True
                            break
                    if matched: break
            mapping_comboboxes[req_col] = cb
        
        if confirm_btn_widget[0]:
            confirm_btn_widget[0].pack(pady=5)
        
        app.update_idletasks()

    def confirm_mapping():
        nonlocal mapping_container
        mapping = {}
        missing = []
        for req_col, cb in mapping_comboboxes.items():
            val = cb.get()
            if not val:
                missing.append(req_col)
            mapping[req_col] = val
        
        if missing:
            CustomMessageBox.showwarning("提示", f"请为以下列选择映射：\n{', '.join(missing)}", parent=app)
            return
        
        try:
            import python_calamine
            engine = 'calamine'
        except ImportError:
            engine = 'openpyxl'
        
        df = pd.read_excel(current_input_path[0], engine=engine)
        df = apply_column_mapping(df, mapping)
        
        required_after = ['会计月', '凭证编号', '一级科目', '借方发生额', '贷方发生额']
        still_missing = [col for col in required_after if col not in df.columns]
        if still_missing:
            CustomMessageBox.showerror("错误", f"映射后仍缺少必要列: {still_missing}", parent=app)
            return
        
        current_df[0] = df
        if mapping_container:
            mapping_container.destroy()
            mapping_container = None
        if confirm_btn_widget[0]:
            confirm_btn_widget[0].pack_forget()
        
        start_processing()

    def start_processing():
        if current_df[0] is None:
            return
        
        df = current_df[0]
        output_path = current_output_path[0]
        
        try:
            anomaly_threshold = float(threshold_entry.get())
        except ValueError:
            anomaly_threshold = 10000.0
        print(f"异常分录筛选阈值设定为: {anomaly_threshold}")
        
        progress_label.configure(text="正在处理...")
        if USE_CUSTOMTKINTER:
            progress_bar.set(0)
        else:
            progress_bar['value'] = 0
        
        def worker():
            try:
                run_processing_pipeline(df, anomaly_threshold, output_path)
                
                def show_success():
                    if USE_CUSTOMTKINTER:
                        progress_bar.set(1.0)
                    else:
                        progress_bar['value'] = 100
                    if os.path.exists(output_path):
                        progress_label.configure(text=f"完成！输出: {os.path.basename(output_path)}")
                        print(f"处理完成！输出文件: {output_path}")
                    else:
                        progress_label.configure(text="完成！但未检测到输出文件")
                
                app.after(0, show_success)
            except Exception as e:
                def show_error_worker():
                    progress_label.configure(text="处理失败")
                    print(f"错误: {str(e)}")
                    traceback.print_exc()
                app.after(0, show_error_worker)
        
        t = threading.Thread(target=worker, daemon=True)
        t.start()

    def select_file():
        nonlocal mapping_container
        input_path = filedialog.askopenfilename(
            title="请选择序时账 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        
        if not input_path: return
        
        current_input_path[0] = input_path
        selected_file_label.configure(text=f"已选择: {os.path.basename(input_path)}")
        print(f"已选择文件: {os.path.basename(input_path)}")
        
        dir_name = os.path.dirname(input_path)
        file_name = os.path.basename(input_path)
        name, ext = os.path.splitext(file_name)
        output_name = f"{name}生成对方科目{ext}"
        output_path = os.path.join(dir_name, output_name)
        current_output_path[0] = output_path
        
        if mapping_container:
            mapping_container.destroy()
            mapping_container = None
        if confirm_btn_widget[0]:
            confirm_btn_widget[0].pack_forget()
        
        need_mapping, file_columns, missing_cols = check_column_mapping_needed(input_path)
        
        if need_mapping:
            progress_label.configure(text="请配置列映射...")
            print(f"检测到缺少必要列: {', '.join(missing_cols)}")
            show_mapping_config(file_columns, missing_cols)
        else:
            try:
                import python_calamine
                engine = 'calamine'
            except ImportError:
                engine = 'openpyxl'
            df = pd.read_excel(input_path, engine=engine)
            current_df[0] = df
            print("文件列名匹配成功，开始处理...")
            start_processing()

    if USE_CUSTOMTKINTER:
        top_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        top_frame.pack(fill="x", pady=(0, 2))
        selected_file_label = ctk.CTkLabel(top_frame, text="请选择序时账Excel文件", font=("微软雅黑", 9), text_color="gray")
        selected_file_label.pack(side="left")
        
        def toggle_theme():
            current = ctk.get_appearance_mode()
            new_mode = "Dark" if current == "Light" else "Light"
            ctk.set_appearance_mode(new_mode)
            theme_btn.configure(text="☀️" if new_mode == "Light" else "🌙")
        
        current_mode = ctk.get_appearance_mode()
        theme_icon = "☀️" if current_mode == "Light" else "🌙"
        theme_btn = ctk.CTkButton(top_frame, text=theme_icon, width=30, height=24, font=("Segoe UI Emoji", 12), command=toggle_theme)
        theme_btn.pack(side="right")
    else:
        selected_file_label = tk.Label(main_frame, text="请选择序时账Excel文件", font=("微软雅黑", 9), fg="gray")
        selected_file_label.pack(pady=(0, 2))

    if USE_CUSTOMTKINTER:
        ctk.CTkButton(main_frame, text="📁 选择Excel文件", font=("微软雅黑", 11), width=160, height=32, command=select_file).pack(pady=3)
    else:
        tk.Button(main_frame, text="📁 选择Excel文件", font=("微软雅黑", 11), width=14, height=1, command=select_file).pack(pady=3)

    threshold_frame = ctk.CTkFrame(main_frame, fg_color="transparent") if USE_CUSTOMTKINTER else tk.Frame(main_frame)
    threshold_frame.pack(pady=3)
    if USE_CUSTOMTKINTER:
        ctk.CTkLabel(threshold_frame, text="异常金额阈值:", font=("微软雅黑", 9)).pack(side="left", padx=(0, 5))
        threshold_entry = ctk.CTkEntry(threshold_frame, width=100, font=("微软雅黑", 10))
        threshold_entry.pack(side="left")
        threshold_entry.insert(0, "10000")
    else:
        tk.Label(threshold_frame, text="异常金额阈值:", font=("微软雅黑", 9)).pack(side="left", padx=(0, 5))
        threshold_entry = tk.Entry(threshold_frame, width=12, font=("微软雅黑", 10))
        threshold_entry.pack(side="left")
        threshold_entry.insert(0, "10000")

    if USE_CUSTOMTKINTER:
        progress_bar = ctk.CTkProgressBar(main_frame, width=300, height=18)
        progress_bar.pack(pady=3)
        progress_bar.set(0)
        progress_label = ctk.CTkLabel(main_frame, text="等待操作...", font=("微软雅黑", 10))
        progress_label.pack()
    else:
        style = ttk.Style()
        style.configure("Custom.Horizontal.TProgressbar", thickness=18)
        progress_bar = ttk.Progressbar(main_frame, length=300, mode='determinate', style="Custom.Horizontal.TProgressbar")
        progress_bar.pack(pady=3)
        progress_bar['value'] = 0
        progress_label = tk.Label(main_frame, text="等待操作...", font=("微软雅黑", 10))
        progress_label.pack()

    if USE_CUSTOMTKINTER:
        confirm_btn_widget[0] = ctk.CTkButton(main_frame, text="确认并开始处理", font=("微软雅黑", 11, "bold"), width=140, height=28, command=confirm_mapping)
    else:
        confirm_btn_widget[0] = tk.Button(main_frame, text="确认并开始处理", font=("微软雅黑", 10), width=12, height=1, command=confirm_mapping)
    confirm_btn_widget[0].pack_forget()

    if USE_CUSTOMTKINTER:
        log_frame = ctk.CTkFrame(main_frame)
        log_frame.pack(fill="both", expand=True, padx=3, pady=5)
        log_text = ctk.CTkTextbox(log_frame, height=120, font=("Consolas", 9))
        log_text.pack(fill="both", expand=True)
    else:
        log_frame = tk.Frame(main_frame)
        log_frame.pack(fill="both", expand=True, padx=3, pady=5)
        log_scroll = tk.Scrollbar(log_frame)
        log_scroll.pack(side="right", fill="y")
        log_text = tk.Text(log_frame, height=8, font=("Consolas", 9), yscrollcommand=log_scroll.set)
        log_text.pack(fill="both", expand=True)
        log_scroll.config(command=log_text.yview)

    def check_queue():
        try:
            while True:
                msg_type, percent, message, phase = GUI_PROGRESS.msg_queue.get_nowait()
                if msg_type == "update":
                    progress_label.configure(text=message)
                    if USE_CUSTOMTKINTER:
                        progress_bar.set(percent / 100.0)
                    else:
                        progress_bar['value'] = percent
                    app.update_idletasks()
                elif msg_type == "stop":
                    pass
        except queue.Empty:
            pass
        
        try:
            while True:
                log_msg = GUI_LOG_QUEUE.get_nowait()
                log_text.configure(state="normal")
                log_text.insert("end", log_msg + "\n")
                log_text.see("end")
                log_text.configure(state="disabled")
        except queue.Empty:
            pass
        
        app.after(100, check_queue)

    GUI_PROGRESS.set_gui_callback(True)
    app.after(100, check_queue)

    if USE_CUSTOMTKINTER:
        ctk.CTkLabel(main_frame, text="v1.8.0", font=("微软雅黑", 8), text_color="gray").pack(side="bottom", pady=(0, 2))
    else:
        tk.Label(main_frame, text="v1.8.0", font=("微软雅黑", 8), fg="gray").pack(side="bottom", pady=(0, 2))

    app.mainloop()
    
    sys.stdout = sys.__stdout__
    sys.stderr = sys.__stderr__

if __name__ == "__main__":
    multiprocessing.freeze_support()
    if len(sys.argv) >= 3:
        generate_contra_account(sys.argv[1], sys.argv[2])
    else:
        run_gui()
