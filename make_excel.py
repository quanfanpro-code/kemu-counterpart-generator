#!/usr/bin/env python3
"""从 DataFrame 生成程序固定使用的深海蓝 Excel 报表。"""

import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════════════════
# 格式常量 — 摩根士丹利标准
# ═══════════════════════════════════════════════════════════════════════════

# 框线（只有上下，无竖线）
THICK = Side(style='medium', color='000000')
DASHED = Side(style='dashed', color='808080')
NO = Side(style=None)

TOP_BORDER = Border(top=THICK, bottom=DASHED, left=NO, right=NO)
MID_BORDER = Border(top=DASHED, bottom=DASHED, left=NO, right=NO)
BOTTOM_BORDER = Border(top=DASHED, bottom=THICK, left=NO, right=NO)

# 表头：深海蓝底白字
HDR_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HDR_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
HDR_ALIGN = Alignment(horizontal='center', vertical='center')

# 数据字体颜色（摩根系：蓝=手动输入，黑=公式结果）
DATA_FONT_BLUE = Font(name='Arial', size=11, color='0000FF')
DATA_FONT_BLACK = Font(name='Arial', size=11, color='000000')

# 合计行样式
SUM_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
SUM_FONT = Font(name='Arial', size=11, bold=True, color='000000')

# 合计行关键词（当最后一行的首列文本匹配时，视为合计行）
SUM_KEYWORDS = ('合计', '小计', '总计', 'sum', 'total', '小計', '合計')

# 数据对齐
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')

# 数字格式
FMT = {
    'money': '#,##0.00',
    'number': '#,##0',
    'pct': '0.00%',
    'date': 'yyyy/mm/dd',
    'text': '@',
    'unknown': '@',
}

# 布局
ROW_HEIGHT = 18
A_COL_WIDTH = 2
MIN_COL_WIDTH = 10
MAX_COL_WIDTH = 50


# ═══════════════════════════════════════════════════════════════════════════
# 列类型推断
# ═══════════════════════════════════════════════════════════════════════════

def _infer_column_type(name: str, series: pd.Series) -> str:
    """根据列名关键词 + 值特征推断列类型。"""
    n = name.lower()

    # ── 关键词匹配（优先级：pct > date > text > money）──
    # "率"优先检测，避免"毛利率"被 money 关键词误匹配
    if re.search(r'(率$|占比|百分比|增长率)', n):
        return 'pct'
    if re.search(r'(日期|时间|年月|期间|年份|月份|date|time)', n):
        return 'date'
    if re.search(r'(编号|id|单号|编码|电话|手机|备注|说明|名称|地址|描述|号码|订单号|负责人|姓名|联系人|部门|岗位)', n):
        return 'text'
    if re.search(r'(金额|价格|收入|成本|费用|毛利额|净利|合计|总额|售价|单价|预算|支出|毛利(?!率))', n):
        return 'money'

    # ── 值特征 ──
    clean = series.dropna()
    if len(clean) == 0:
        return 'unknown'

    try:
        if pd.api.types.is_datetime64_any_dtype(clean):
            return 'date'

        if pd.api.types.is_numeric_dtype(clean):
            sample = clean.head(20)
            if pd.api.types.is_float_dtype(sample):
                # 0~1 之间的浮点 → 百分比
                if (sample >= 0).all() and (sample <= 1).all() and sample.nunique() > 2:
                    return 'pct'
                return 'money'
            return 'number'

        # 长数字字符串 → text
        str_sample = clean.astype(str).head(20)
        if str_sample.str.match(r'^\d{12,}$').any():
            return 'text'
    except Exception:
        pass

    return 'unknown'


# ═══════════════════════════════════════════════════════════════════════════
# 列宽估算
# ═══════════════════════════════════════════════════════════════════════════

def _char_width(text: str) -> int:
    """中文字符计2，拉丁/数字计1。"""
    return sum(2 if ord(ch) > 127 else 1 for ch in str(text))


def _estimate_column_width(header: str, series: pd.Series) -> int:
    """估算列宽：取表头和数据中最大字符宽度 + padding。"""
    max_w = _char_width(header)
    for v in series.dropna():
        try:
            w = _char_width(str(v))
            if w > max_w:
                max_w = w
        except Exception:
            pass
    return max(MIN_COL_WIDTH, min(max_w + 2, MAX_COL_WIDTH))


# ═══════════════════════════════════════════════════════════════════════════
# 格式引擎
# ═══════════════════════════════════════════════════════════════════════════

def _apply_styles(ws, df: pd.DataFrame):
    """对已写入数据的工作表应用摩根系格式。"""
    nrows = 1 + len(df)   # 表头 + 数据行
    ncols = len(df.columns)
    start_col = 2         # B 列

    # 1. 行高 + A列宽
    for r in range(1, nrows + 1):
        ws.row_dimensions[r].height = ROW_HEIGHT
    ws.column_dimensions['A'].width = A_COL_WIDTH

    # 2. 列宽估算 + 类型推断（一次扫描）
    col_types = {}
    for ci in range(ncols):
        col_name = df.columns[ci]
        col_letter = get_column_letter(ci + start_col)
        col_type = _infer_column_type(str(col_name), df.iloc[:, ci])
        col_types[ci] = col_type

        est = _estimate_column_width(str(col_name), df.iloc[:, ci])
        ws.column_dimensions[col_letter].width = est

    # 3. 表头（第1行）
    for ci in range(ncols):
        cell = ws.cell(row=1, column=ci + start_col)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = HDR_ALIGN

    # 4. 数据行（统一 Arial 11 + 数字格式 + 对齐 + 字体颜色）
    for ci in range(ncols):
        ct = col_types.get(ci, 'unknown')
        is_num = ct in ('money', 'number', 'pct')
        align = RIGHT_ALIGN if is_num else LEFT_ALIGN
        nf = FMT.get(ct)

        for ri in range(2, nrows + 1):
            cell = ws.cell(row=ri, column=ci + start_col)
            # 所有单元格统一 Arial 11
            cell.font = DATA_FONT_BLUE if is_num else Font(name='Arial', size=11)
            cell.alignment = align
            if nf:
                cell.number_format = nf

    # 5. 合计行检测与特殊格式（最后一行的首列含合计关键词）
    last_val = ws.cell(row=nrows, column=start_col).value
    is_summary = (
        nrows > 1
        and last_val is not None
        and any(kw in str(last_val) for kw in SUM_KEYWORDS)
    )
    if is_summary:
        for ci in range(ncols):
            cell = ws.cell(row=nrows, column=ci + start_col)
            cell.font = SUM_FONT
            cell.fill = SUM_FILL

    # 6. 边框（无竖线，上下粗中间虚线）
    for ci in range(ncols):
        col = ci + start_col
        for ri in range(1, nrows + 1):
            cell = ws.cell(row=ri, column=col)
            if ri == 1:
                cell.border = TOP_BORDER
            elif ri == nrows:
                cell.border = BOTTOM_BORDER
            else:
                cell.border = MID_BORDER

    # 6. 右侧空白列（宽 3）
    last_data_col = start_col + ncols - 1
    right_col = get_column_letter(last_data_col + 1)
    ws.column_dimensions[right_col].width = 3

    # 7. 网格线 + 冻结
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A2'


# ═══════════════════════════════════════════════════════════════════════════
# 写数据
# ═══════════════════════════════════════════════════════════════════════════

def _clean_cell(value):
    """NaN/NA/NaT 无法写入 Excel，统一置空。"""
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    return value


def _write_data(ws, df: pd.DataFrame):
    """从 B2 开始写入（第1行 = 表头，A列留空）。"""
    for ci, col_name in enumerate(df.columns):
        ws.cell(row=1, column=ci + 2, value=col_name)
    # itertuples 比 iterrows 快一个数量级
    for ri, row in enumerate(df.itertuples(index=False, name=None), start=2):
        for ci, value in enumerate(row):
            ws.cell(row=ri, column=ci + 2, value=_clean_cell(value))


# ═══════════════════════════════════════════════════════════════════════════
# 主入口
# ═══════════════════════════════════════════════════════════════════════════

def make_excel(data, output_path: str, sheet_name: str = 'Sheet1') -> str:
    """生成摩根系标准格式 Excel。

    参数
    ----
    data : DataFrame | list[tuple[str, DataFrame]]
        单 sheet 传 df，多 sheet 传 [(名称, df), ...]。
    output_path : str
        输出文件路径。
    sheet_name : str
        单 sheet 模式的工作表名称。
    返回
    ----
    str : 输出文件的绝对路径。
    """
    # 统一为多 sheet 格式
    if isinstance(data, pd.DataFrame):
        sheets = [(sheet_name, data)]
    else:
        sheets = list(data)

    wb = Workbook()
    for name in list(wb.sheetnames):
        del wb[name]

    for idx, (name, df) in enumerate(sheets):
        if df is None or (hasattr(df, 'empty') and df.empty):
            continue
        safe_name = re.sub(r'[\\/*?:\[\]]', '_', str(name))[:31]
        ws = wb.create_sheet(title=safe_name, index=idx)
        _write_data(ws, df)
        _apply_styles(ws, df)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return str(out.absolute())
