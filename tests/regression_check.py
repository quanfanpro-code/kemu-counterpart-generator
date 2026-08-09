# -*- coding: utf-8 -*-
"""生成修改前基线，并验证工作簿除首行水平对齐外完全一致。"""
from __future__ import annotations

import argparse
import os
import subprocess
import sys
import tempfile
from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_CLI = ROOT / "cli" / "序时账分析器命令行版.py"
OUTPUT_NAMES = ("main-full.xlsx", "cli-summary.xlsx", "cli-full.xlsx")


def _write_inputs(directory: Path) -> tuple[Path, Path]:
    rows = [
        {
            "账套": "测试账套",
            "记账时间": "2026-01-05",
            "会计月": "2026-01",
            "凭证种类": "记",
            "凭证编号": "1",
            "编号": "1-1",
            "业务说明": "收到货款",
            "一级编号": "1002",
            "一级科目": "银行存款",
            "科目编号": "100201",
            "科目名称": "基本户",
            "借方发生额": 1000,
            "贷方发生额": 0,
        },
        {
            "账套": "测试账套",
            "记账时间": "2026-01-05",
            "会计月": "2026-01",
            "凭证种类": "记",
            "凭证编号": "1",
            "编号": "1-2",
            "业务说明": "收到货款",
            "一级编号": "6001",
            "一级科目": "主营业务收入",
            "科目编号": "600101",
            "科目名称": "商品销售收入",
            "借方发生额": 0,
            "贷方发生额": 1000,
        },
        {
            "账套": "测试账套",
            "记账时间": "2026-01-06",
            "会计月": "2026-01",
            "凭证种类": "记",
            "凭证编号": "2",
            "编号": "2-1",
            "业务说明": "支付费用",
            "一级编号": "6602",
            "一级科目": "管理费用",
            "科目编号": "660201",
            "科目名称": "办公费",
            "借方发生额": 200,
            "贷方发生额": 0,
        },
        {
            "账套": "测试账套",
            "记账时间": "2026-01-06",
            "会计月": "2026-01",
            "凭证种类": "记",
            "凭证编号": "2",
            "编号": "2-2",
            "业务说明": "支付费用",
            "一级编号": "1002",
            "一级科目": "银行存款",
            "科目编号": "100201",
            "科目名称": "基本户",
            "借方发生额": 0,
            "贷方发生额": 200,
        },
    ]
    df = pd.DataFrame(rows)
    valid = directory / "representative-input.xlsx"
    invalid = directory / "missing-debit-input.xlsx"
    df.to_excel(valid, index=False)
    df.drop(columns=["借方发生额"]).to_excel(invalid, index=False)
    return valid, invalid


def _run(command: list[str]) -> subprocess.CompletedProcess[str]:
    env = os.environ.copy()
    env["PYTHONUTF8"] = "1"
    return subprocess.run(
        command,
        cwd=ROOT,
        env=env,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=180,
    )


def _require_exit(command: list[str], expected: int) -> None:
    result = _run(command)
    if result.returncode != expected:
        raise AssertionError(
            f"退出码不符：期望 {expected}，实际 {result.returncode}\n"
            f"命令：{command}\n标准输出：{result.stdout}\n错误输出：{result.stderr}"
        )


def _source_commands(valid: Path, output_dir: Path) -> dict[str, list[str]]:
    return {
        "main-full.xlsx": [
            sys.executable, "-B", "-X", "utf8", str(ROOT / "main.py"),
            str(valid), str(output_dir / "main-full.xlsx"), "--no-gui",
            "--log-level", "ERROR",
        ],
        "cli-summary.xlsx": [
            sys.executable, "-B", "-X", "utf8", str(SOURCE_CLI),
            str(valid), str(output_dir / "cli-summary.xlsx"),
            "--mode", "summary", "--log-level", "ERROR",
        ],
        "cli-full.xlsx": [
            sys.executable, "-B", "-X", "utf8", str(SOURCE_CLI),
            str(valid), str(output_dir / "cli-full.xlsx"),
            "--mode", "full", "--log-level", "ERROR",
        ],
    }


def _exe_commands(exe: Path, valid: Path, output_dir: Path) -> dict[str, list[str]]:
    return {
        "cli-summary.xlsx": [
            str(exe), str(valid), str(output_dir / "cli-summary.xlsx"),
            "--mode", "summary", "--log-level", "ERROR",
        ],
        "cli-full.xlsx": [
            str(exe), str(valid), str(output_dir / "cli-full.xlsx"),
            "--mode", "full", "--log-level", "ERROR",
        ],
    }


def _invalid_commands(runner: str, invalid: Path, output_dir: Path,
                      exe: Path | None) -> list[list[str]]:
    if runner == "exe":
        if exe is None:
            raise AssertionError("EXE 验证缺少 --exe 参数")
        return [[str(exe), str(invalid), str(output_dir / "invalid.xlsx"),
                 "--log-level", "ERROR"]]
    return [
        [sys.executable, "-B", "-X", "utf8", str(ROOT / "main.py"),
         str(invalid), str(output_dir / "invalid-main.xlsx"), "--no-gui",
         "--log-level", "ERROR"],
        [sys.executable, "-B", "-X", "utf8", str(SOURCE_CLI),
         str(invalid), str(output_dir / "invalid-cli.xlsx"),
         "--log-level", "ERROR"],
    ]


def _dimension_state(dimensions) -> dict[str, tuple]:
    return {
        str(key): (
            value.height if hasattr(value, "height") else value.width,
            value.hidden,
            value.outlineLevel,
            value.collapsed,
        )
        for key, value in dimensions.items()
    }


def _alignment_without_horizontal(cell) -> tuple:
    alignment = cell.alignment
    return (
        alignment.vertical,
        alignment.textRotation,
        alignment.wrapText,
        alignment.shrinkToFit,
        alignment.indent,
        alignment.relativeIndent,
        alignment.justifyLastLine,
        alignment.readingOrder,
    )


def assert_headers_centered(path: Path) -> None:
    workbook = load_workbook(path, data_only=False)
    for worksheet in workbook.worksheets:
        for cell in worksheet[1]:
            if cell.value is not None:
                assert cell.alignment.horizontal == "center", (
                    path, worksheet.title, cell.coordinate,
                    cell.alignment.horizontal,
                )


def assert_workbooks_equal_except_header_alignment(before: Path, after: Path) -> None:
    old = load_workbook(before, data_only=False)
    new = load_workbook(after, data_only=False)
    assert old.sheetnames == new.sheetnames
    for old_ws, new_ws in zip(old.worksheets, new.worksheets):
        assert old_ws.max_row == new_ws.max_row
        assert old_ws.max_column == new_ws.max_column
        assert old_ws.sheet_state == new_ws.sheet_state
        assert old_ws.freeze_panes == new_ws.freeze_panes
        assert old_ws.sheet_view.showGridLines == new_ws.sheet_view.showGridLines
        assert old_ws.auto_filter.ref == new_ws.auto_filter.ref
        assert list(old_ws.merged_cells.ranges) == list(new_ws.merged_cells.ranges)
        assert len(old_ws._charts) == len(new_ws._charts)
        assert _dimension_state(old_ws.row_dimensions) == _dimension_state(new_ws.row_dimensions)
        assert _dimension_state(old_ws.column_dimensions) == _dimension_state(new_ws.column_dimensions)
        for row in old_ws.iter_rows(
            min_row=1,
            max_row=old_ws.max_row,
            min_col=1,
            max_col=old_ws.max_column,
        ):
            for old_cell in row:
                new_cell = new_ws[old_cell.coordinate]
                assert old_cell.value == new_cell.value
                assert old_cell.data_type == new_cell.data_type
                assert old_cell.number_format == new_cell.number_format
                assert copy(old_cell.font) == copy(new_cell.font)
                assert copy(old_cell.fill) == copy(new_cell.fill)
                assert copy(old_cell.border) == copy(new_cell.border)
                assert copy(old_cell.protection) == copy(new_cell.protection)
                assert old_cell.hyperlink == new_cell.hyperlink
                assert old_cell.comment == new_cell.comment
                if old_cell.row == 1 and old_cell.value is not None:
                    assert new_cell.alignment.horizontal == "center"
                    assert _alignment_without_horizontal(old_cell) == _alignment_without_horizontal(new_cell)
                else:
                    assert copy(old_cell.alignment) == copy(new_cell.alignment)


def _capture(directory: Path) -> None:
    directory.mkdir(parents=True, exist_ok=True)
    valid, invalid = _write_inputs(directory)
    for name, command in _source_commands(valid, directory).items():
        _require_exit(command, 0)
        assert (directory / name).is_file()
    for command in _invalid_commands("source", invalid, directory, None):
        _require_exit(command, 1)
    print("BASELINE_CAPTURED outputs=3 error_exit=1")


def _verify(directory: Path, runner: str, exe: Path | None,
            expect_centered: bool) -> None:
    valid = directory / "representative-input.xlsx"
    invalid = directory / "missing-debit-input.xlsx"
    for required in (valid, invalid, *(directory / name for name in OUTPUT_NAMES)):
        if not required.is_file():
            raise AssertionError(f"基线文件不存在：{required}")
    if runner == "exe" and (exe is None or not exe.is_file()):
        raise AssertionError(f"EXE 不存在：{exe}")

    with tempfile.TemporaryDirectory(prefix="ledger-regression-") as temp:
        output_dir = Path(temp)
        commands = (_source_commands(valid, output_dir) if runner == "source"
                    else _exe_commands(exe, valid, output_dir))
        for name, command in commands.items():
            _require_exit(command, 0)
            current = output_dir / name
            assert_workbooks_equal_except_header_alignment(directory / name, current)
            if expect_centered:
                assert_headers_centered(current)
        for command in _invalid_commands(runner, invalid, output_dir, exe):
            _require_exit(command, 1)
    print(
        f"REGRESSION_OK outputs={len(commands)} "
        "allowed_differences=header_horizontal_only error_exit=1"
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    action = parser.add_mutually_exclusive_group(required=True)
    action.add_argument("--capture", type=Path)
    action.add_argument("--verify", type=Path)
    parser.add_argument("--runner", choices=("source", "exe"), required=True)
    parser.add_argument("--exe", type=Path)
    parser.add_argument("--expect-centered", action="store_true")
    args = parser.parse_args()

    if args.capture:
        if args.runner != "source":
            parser.error("捕获基线只允许使用 source")
        _capture(args.capture.resolve())
    else:
        _verify(args.verify.resolve(), args.runner,
                args.exe.resolve() if args.exe else None,
                args.expect_centered)


if __name__ == "__main__":
    main()
