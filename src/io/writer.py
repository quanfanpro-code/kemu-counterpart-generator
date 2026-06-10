# -*- coding: utf-8 -*-
"""数据输出模块 — 摩根系 deep-navy 标准 + 班福图表"""
from typing import List, Dict, Callable, Optional

import pandas as pd

try:
    from openpyxl import load_workbook
    from openpyxl.chart import LineChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    load_workbook = LineChart = Reference = None

try:
    from make_excel import make_excel
    MAKE_EXCEL_AVAILABLE = True
except ImportError:
    MAKE_EXCEL_AVAILABLE = False

from ..utils.logger import logger


def save_output_file(output_path: str, df: pd.DataFrame, out_df: pd.DataFrame,
                     anomaly_df: pd.DataFrame, aggregated_patterns: pd.DataFrame,
                     stats_df: pd.DataFrame, failed_groups: List[Dict],
                     progress_callback: Optional[Callable] = None):
    """
    保存结果到 Excel（摩根系 deep-navy 标准格式）。

    :param output_path: 输出文件路径
    :param df: 原始数据
    :param out_df: 生成结果
    :param anomaly_df: 异常分录明细
    :param aggregated_patterns: 异常分录汇总
    :param stats_df: 班福分析数据
    :param failed_groups: 失败分组列表
    :param progress_callback: 进度回调函数 callback(percent, message, phase)
    """
    logger.info(f"正在保存文件到: {output_path}")
    if progress_callback:
        progress_callback(92, "正在保存Excel文件...", "文件输出")

    if not OPENPYXL_AVAILABLE:
        logger.error("错误: openpyxl库未安装，无法生成Excel文件。请运行: pip install openpyxl")
        if progress_callback:
            progress_callback(0, "保存失败", "错误")
        return

    try:
        logger.info("正在写入Excel文件(包含多个Sheet及图表)...")

        # ── 整理列顺序（账套列如果存在则放在最前面）──
        ledger_col = df.attrs.get('ledger_column', None)
        cols = []
        if ledger_col and ledger_col in out_df.columns:
            cols.append(ledger_col)
        cols.extend(['记账时间', '会计月', '凭证种类', '凭证编号', '编号', '业务说明', '一级编号', '一级科目', '科目编号', '科目名称', '借方发生额', '贷方发生额', '对方科目', '匹配类型'])
        for c in df.columns:
            if c not in cols and c != '首位数': cols.append(c)
        final_cols = [c for c in cols if c in out_df.columns]
        out_df = out_df[final_cols]

        # ── 构建多 Sheet 数据 ──
        sheets = [
            ('生成结果', out_df),
            ('原始数据', df),
        ]

        if not anomaly_df.empty:
            sheets.append(('异常分录明细', anomaly_df))
        else:
            sheets.append(('异常分录明细', pd.DataFrame({"提示": ["未检测到符合阈值的异常分录"]})))

        if isinstance(aggregated_patterns, pd.DataFrame) and not aggregated_patterns.empty:
            sheets.append(('异常分录', aggregated_patterns))
        else:
            sheets.append(('异常分录', pd.DataFrame(columns=['借方科目', '贷方科目', '业务描述', '金额', '凭证编号', '异常类型编号'])))

        if failed_groups:
            sheets.append(('失败分组', pd.DataFrame(failed_groups)))

        sheets.append(('班福分析', stats_df))

        # ── 使用 make_excel 生成摩根系标准格式 ──
        if MAKE_EXCEL_AVAILABLE:
            make_excel(sheets, output_path, theme='deep-navy')
        else:
            # 回退：无 make_excel 时用裸 to_excel
            logger.warning("make_excel 不可用，使用默认格式输出")
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for name, sheet_df in sheets:
                    sheet_df.to_excel(writer, sheet_name=name, index=False)

        # ── 添加班福图表（make_excel 从 B2 写，坐标需偏移 +1）──
        wb = load_workbook(output_path)
        stats_ws = wb['班福分析']

        # make_excel 从 B 列(col=2) 第 1 行写表头，数据从第 2 行开始
        # 班福分析表：首位数在 col=2, 班福频率在 col=4, 实际频率在 col=5
        # 图表数据引用：表头在第 1 行，数据在第 2~10 行（首位数 1-9）
        chart = LineChart()
        chart.title = "班福定律分析 - 首位数分布"
        chart.style = 13
        chart.y_axis.title = "比率"
        chart.x_axis.title = "首位数"
        cats = Reference(stats_ws, min_col=2, min_row=2, max_row=10)
        data = Reference(stats_ws, min_col=4, min_row=1, max_row=10, max_col=5)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        stats_ws.add_chart(chart, "H2")

        wb.save(output_path)

        logger.info("完成。")
        if progress_callback:
            progress_callback(100, "处理完成", "完成")
    except Exception as e:
        logger.error(f"保存文件失败: {e}")
        if progress_callback:
            progress_callback(0, "保存失败", "错误")
